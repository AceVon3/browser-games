"""Cross-check Idaho scraped data against AM Best Report.pdf ground truth.

Constraints discovered during reconciliation:
  - PDF SERFF tracking numbers are MASKED as '*****', so SERFF-based matching is impossible
  - PDF dates are effective dates; our dataset has submission_date / disposition_date
  - id_final.xlsx rate-value columns are almost entirely None (values live in PDF parser
    output, not the search dataset), so per-row value comparison can only be done where
    rate_changes.xlsx has extracted values

Matching strategy:
  - For each PDF filing, find rows whose company_name matches exactly (or per a synonym
    map for renames) AND whose disposition_date precedes the PDF effective date by
    <= MAX_DISP_TO_EFF_DAYS.
  - Each of our rows may be used at most once (greedy: closest disposition_date wins).
  - Bucket 3 is restricted to target TOIs only (19.x PPA, 04.x Homeowners, 03.x Farmowners).
"""
from __future__ import annotations
import sys
from datetime import datetime, date
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")
from src.config import OUTPUT_DIR

MAX_DISP_TO_EFF_DAYS = 270   # disposition should precede effective date by up to ~9mo
TARGET_TOI_PREFIXES = ("19.0", "04.0", "03.0")  # PPA, Homeowners, Farmowners

# Ground truth from AM Best Report.pdf (Idaho, target carriers, PPA only)
# Tuple: (carrier_group, company, effective_mmddyy, indicated_change, rate_impact, policyholders, page_section)
PDF_TARGET_FILINGS = [
    ("GEICO/Berkshire", "GEICO Casualty Company",                         "04/03/26", "-12.400", "-7.500", 9434,   "p1"),
    ("GEICO/Berkshire", "GEICO Marine Insurance Company",                 "04/03/26", "-11.900", "-7.500", 1447,   "p2"),
    ("State Farm",      "State Farm Mutual Automobile Insurance Company", "01/02/26", "-2.600",  "-9.700", 360274, "p2"),
    ("State Farm",      "State Farm Fire and Casualty Company",           "01/02/26", "15.900",  "-2.100", 20679,  "p3"),
    ("State Farm",      "MGA Insurance Company, Inc.",                    "01/23/26", "3.700",   "-0.100", 2534,   "p6"),
    ("Liberty Mutual",  "American Economy Insurance Company",             "03/21/26", "-2.200",  "0.000",  21727,  "p5"),
    ("Allstate",        "Integon National Insurance Company",             "11/03/25", "0.000",   "-9.600", 61,     "p16"),
    ("Travelers",       "The Standard Fire Insurance Company",            "09/26/25", "-1.800",  "-1.600", 17646,  "p17"),
    ("Allstate",        "Allstate North American Insurance Company",      "09/23/25", "",        "-3.900", 6146,   "p19"),
    ("Allstate",        "Allstate North American Insurance Company",      "02/10/26", "",        "",       0,      "p19_NA"),
    ("Travelers",       "The Standard Fire Insurance Company",            "11/16/25", "",        "0.000",  17676,  "p20"),
    ("Allstate",        "Allstate Indemnity Company",                     "09/29/25", "0.000",   "0.000",  0,      "p21"),
    ("Allstate",        "Allstate Insurance Company",                     "09/29/25", "0.000",   "0.000",  0,      "p21"),
    ("Allstate",        "Allstate Property and Casualty Insurance Company","09/29/25","0.000",   "0.000",  0,      "p21"),
    ("Allstate",        "Integon Indemnity Corporation",                  "05/29/25", "7.700",   "0.500",  5555,   "p23"),
    ("Allstate",        "Encompass Indemnity Company",                    "09/01/25", "10.400",  "10.400", 785,    "p24"),
    ("Allstate",        "Allstate Fire and Casualty Insurance Company",   "06/23/25", "6.000",   "6.000",  8410,   "p29"),
    ("Allstate",        "Esurance Property and Casualty Insurance Company","06/05/25","0.000",   "0.000",  0,      "p31"),
    ("Allstate",        "Allstate Fire and Casualty Insurance Company",   "05/19/25", "0.000",   "0.000",  0,      "p31"),
    ("Allstate",        "Encompass Indemnity Company",                    "06/23/25", "0.000",   "0.000",  0,      "p34"),
    ("Allstate",        "Integon National Insurance Company",             "04/28/25", "0.000",   "0.000",  12,     "p34"),
    ("Allstate",        "Allstate North American Insurance Company",      "07/08/25", "",        "1.000",  158,    "p39"),
    ("Liberty Mutual",  "Safeco Insurance Company of Illinois",           "05/12/25", "-12.700", "-3.000", 48818,  "p39"),
    ("Liberty Mutual",  "Liberty Mutual Insurance Company",               "07/21/25", "0.000",   "0.000",  0,      "p40"),
    ("Liberty Mutual",  "Liberty Mutual Personal Insurance Company",      "07/21/25", "1.300",   "0.000",  2904,   "p40"),
]

GROUP_KEYWORDS = {
    "State Farm":      ["state farm", "mga insurance"],
    "GEICO":           ["geico"],
    "GEICO/Berkshire": ["geico"],
    "Progressive":     ["progressive"],
    "Allstate":        ["allstate", "encompass", "esurance", "integon", "north american insurance"],
    "Travelers":       ["travelers", "standard fire"],
    "Liberty Mutual":  ["liberty mutual", "safeco", "american economy"],
}


def carrier_group(name: str) -> str | None:
    if not name:
        return None
    n = name.lower()
    if "state farm" in n or "mga insurance" in n:
        return "State Farm"
    if "geico" in n:
        return "GEICO"
    if "progressive" in n:
        return "Progressive"
    if any(k in n for k in ["allstate", "encompass", "esurance", "integon", "north american insurance"]):
        return "Allstate"
    if "travelers" in n or "standard fire" in n:
        return "Travelers"
    if any(k in n for k in ["liberty mutual", "safeco", "american economy"]):
        return "Liberty Mutual"
    return None


def parse_eff(s: str) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%m/%d/%y").date()
    except ValueError:
        return None


def to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def load_id_final() -> list[dict]:
    p = OUTPUT_DIR / "id_final.xlsx"
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, r))
        rows.append(d)
    return rows


def load_rate_changes() -> dict[str, dict]:
    """Map serff -> row for Idaho rows in rate_changes.xlsx (or csv fallback)."""
    out = {}
    p = OUTPUT_DIR / "rate_changes.xlsx"
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb["Rate Changes"]
        hdr = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(hdr, r))
            if (d.get("state") or "").upper() != "ID":
                continue
            serff = d.get("serff_tracking_number")
            if serff:
                out.setdefault(serff, []).append(d)
        return out
    except PermissionError:
        print("  (rate_changes.xlsx locked — falling back to rate_changes.csv)")
    import csv as _csv
    csv_p = OUTPUT_DIR / "rate_changes.csv"
    with open(csv_p, encoding="utf-8") as fh:
        for d in _csv.DictReader(fh):
            if (d.get("state") or "").upper() != "ID":
                continue
            serff = d.get("serff_tracking_number")
            if serff:
                out.setdefault(serff, []).append(d)
    return out


def main():
    all_rows = load_id_final()
    rc_map = load_rate_changes()
    print(f"id_final.xlsx Idaho rows:      {len(all_rows)}")
    print(f"rate_changes.* Idaho serffs:   {len(rc_map)}")

    # Filter to target groups + target TOIs
    target_rows = []
    for d in all_rows:
        g = carrier_group(d.get("company_name") or "")
        if not g:
            continue
        toi = d.get("type_of_insurance") or ""
        if not any(toi.startswith(t) for t in TARGET_TOI_PREFIXES):
            continue
        d["_group"] = g
        d["_disp"] = to_date(d.get("disposition_date"))
        d["_sub"] = to_date(d.get("submission_date"))
        target_rows.append(d)
    print(f"Target-group + target-TOI rows:{len(target_rows)}")
    print(f"PDF target-carrier filings:    {len(PDF_TARGET_FILINGS)}\n")

    # Counts per (group, company) on both sides
    from collections import Counter
    pdf_counts = Counter((g, c) for g, c, *_ in PDF_TARGET_FILINGS)
    ours_counts = Counter((d["_group"], d.get("company_name")) for d in target_rows)

    # ------------- Match each PDF filing to our rows -------------
    # strategy: exact company name match + disposition_date before PDF eff_date, within window
    used_serffs: set[str] = set()
    in_both = []
    not_matched_pdf = []
    for pdf_f in PDF_TARGET_FILINGS:
        group, company, eff_str, pdf_ind, pdf_imp, pdf_ph, page = pdf_f
        eff = parse_eff(eff_str)
        candidates = []
        for d in target_rows:
            if d["serff_tracking_number"] in used_serffs:
                continue
            if (d.get("company_name") or "").strip().lower() != company.strip().lower():
                continue
            disp = d["_disp"] or d["_sub"]
            if eff and disp:
                gap = (eff - disp).days
                if gap < 0 or gap > MAX_DISP_TO_EFF_DAYS:
                    continue
                candidates.append((gap, d))
            else:
                candidates.append((999, d))
        if not candidates:
            not_matched_pdf.append(pdf_f)
            continue
        candidates.sort(key=lambda x: x[0])
        best = candidates[0][1]
        used_serffs.add(best["serff_tracking_number"])
        in_both.append((pdf_f, best))

    # ------------- BUCKET 1: IN BOTH -------------
    print("=" * 78)
    print(f"BUCKET 1 (GREEN): IN BOTH  —  {len(in_both)} of {len(PDF_TARGET_FILINGS)}")
    print("=" * 78)
    value_mismatches = []
    value_unknown = 0
    for pdf_f, row in in_both:
        group, company, eff_str, pdf_ind, pdf_imp, pdf_ph, page = pdf_f
        serff = row["serff_tracking_number"]
        rc_rows = rc_map.get(serff, [])
        rc = rc_rows[0] if rc_rows else {}
        our_imp = rc.get("rate_effect_value") or rc.get("rc_rate_effect_value")
        our_ind = rc.get("overall_indicated_change") or rc.get("rc_overall_indicated_change")
        our_ph  = rc.get("policyholders_affected") or rc.get("rc_policyholders_affected")

        def norm(v):
            if v in (None, "", "Data Not Available"):
                return None
            s = str(v).replace("%", "").replace("+", "").replace(",", "").strip()
            try:
                return round(float(s), 3)
            except ValueError:
                return None

        flags = []
        if pdf_imp and norm(our_imp) is not None and norm(our_imp) != norm(pdf_imp):
            flags.append(f"impact: ours={our_imp} pdf={pdf_imp}%")
        if pdf_ind and norm(our_ind) is not None and norm(our_ind) != norm(pdf_ind):
            flags.append(f"indicated: ours={our_ind} pdf={pdf_ind}%")
        if pdf_ph and norm(our_ph) is not None and norm(our_ph) != norm(pdf_ph):
            flags.append(f"policyholders: ours={our_ph} pdf={pdf_ph}")

        if rc_rows:
            status = "[MATCH]" if not flags else "[VALUE MISMATCH]"
        else:
            status = "[no rate_changes row — cannot compare values]"
            value_unknown += 1

        print(f"\n  [{group}] {company}")
        print(f"    our serff={serff}  disp={row['_disp']}  sub={row['_sub']}")
        print(f"    pdf eff={eff_str}  pdf_impact={pdf_imp}  pdf_ind={pdf_ind}  pdf_ph={pdf_ph}")
        print(f"    {status}")
        for fl in flags:
            print(f"      !! {fl}")
            value_mismatches.append((serff, fl))

    # ------------- BUCKET 2: IN PDF BUT NOT OURS -------------
    print("\n" + "=" * 78)
    print(f"BUCKET 2 (RED): IN PDF BUT NOT OURS  —  {len(not_matched_pdf)}")
    print("=" * 78)
    for pdf_f in not_matched_pdf:
        group, company, eff_str, pdf_ind, pdf_imp, pdf_ph, page = pdf_f
        print(f"\n  [{group}] {company}")
        print(f"    pdf eff={eff_str}  indicated={pdf_ind}%  impact={pdf_imp}%  policyholders={pdf_ph}  ({page})")
        # Explain why it wasn't matched
        same_co_rows = [d for d in target_rows if (d.get("company_name") or "").lower() == company.lower()]
        if not same_co_rows:
            print(f"    Reason: we have NO Idaho rows for this company (target TOIs).")
            # Check if we have out-of-TOI rows for this company
            out_of_toi = [d for d in all_rows if (d.get("company_name") or "").lower() == company.lower()]
            if out_of_toi:
                print(f"    (We do have {len(out_of_toi)} Idaho rows for this company in other TOIs)")
        else:
            eff = parse_eff(eff_str)
            print(f"    Reason: we have {len(same_co_rows)} rows for this company but none within"
                  f" {MAX_DISP_TO_EFF_DAYS}d before eff={eff_str}:")
            for d in same_co_rows[:4]:
                used = "  USED-BY-OTHER-MATCH" if d["serff_tracking_number"] in used_serffs else ""
                print(f"      {d['serff_tracking_number']}  disp={d['_disp']}  sub={d['_sub']}{used}")

    # ------------- BUCKET 3: IN OURS BUT NOT PDF (target TOIs only) -------------
    print("\n" + "=" * 78)
    extras = [d for d in target_rows if d["serff_tracking_number"] not in used_serffs]
    print(f"BUCKET 3 (YELLOW): IN OURS BUT NOT PDF  —  {len(extras)}")
    print("(target TOIs only: 19.x PPA, 04.x Homeowners, 03.x Farmowners)")
    print("=" * 78)
    # Group by carrier group
    by_group: dict[str, list] = {}
    for d in extras:
        by_group.setdefault(d["_group"], []).append(d)
    for g in sorted(by_group):
        rows = by_group[g]
        print(f"\n  [{g}] — {len(rows)} extra")
        for d in sorted(rows, key=lambda x: (x.get("company_name") or "", x["_sub"] or date.min)):
            toi = (d.get("type_of_insurance") or "")[:20]
            status = d.get("disposition_status") or d.get("filing_status") or ""
            print(f"    {d['serff_tracking_number']:20s}  disp={d['_disp']}  "
                  f"{(d.get('company_name') or '')[:42]:42s}  toi={toi:20s}  status={status}")

    # ------------- SIDE-BY-SIDE COMPANY COUNTS -------------
    print("\n" + "=" * 78)
    print("SIDE-BY-SIDE COMPANY COUNTS")
    print("=" * 78)
    print(f"  {'carrier':16s} {'company':55s} {'PDF':>4s} {'ours':>5s}")
    companies = sorted(set(list(pdf_counts) + list(ours_counts)), key=lambda x: (x[0], x[1] or ""))
    for key in companies:
        g, c = key
        pc = pdf_counts.get(key, 0)
        oc = ours_counts.get(key, 0)
        marker = ""
        if pc > oc: marker = "  <-- PDF > ours"
        elif oc > pc: marker = "  <-- ours > PDF"
        print(f"  {g:16s} {(c or '')[:55]:55s} {pc:>4d} {oc:>5d}{marker}")

    # ------------- SUMMARY -------------
    print("\n" + "=" * 78)
    print("SUMMARY")
    print("=" * 78)
    print(f"  PDF target-carrier filings (Idaho PPA):     {len(PDF_TARGET_FILINGS)}")
    print(f"    Matched in our dataset:                   {len(in_both)}")
    print(f"    Unmatched (potential misses):             {len(not_matched_pdf)}")
    print(f"  Rows in ours (target TOIs) not in PDF:      {len(extras)}")
    print(f"  Matched with rate_changes comparison data:  {len(in_both) - value_unknown}")
    print(f"  Value mismatches flagged:                   {len(value_mismatches)}")
    print()
    print("  CAVEATS:")
    print("    * PDF SERFFs are masked as *****, so matching is by")
    print("      carrier_group + exact company_name + (disp_date <= eff_date <= disp+270d).")
    print("    * id_final.xlsx rate-value columns are empty for Idaho, so value comparison")
    print("      only happens for the subset of rows present in rate_changes.xlsx.")
    print("    * PDF report covers PPA only for Idaho — no Homeowners or Progressive.")
    print("    * Bucket 3 is restricted to target TOIs (19.x PPA, 04.x HO, 03.x Farmowners)")
    print("      so out-of-scope lines (Other Liability, Workers Comp, etc.) aren't listed.")


if __name__ == "__main__":
    main()
