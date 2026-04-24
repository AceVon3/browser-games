"""Merge per-carrier OR search checkpoints into or_all_companies_search.xlsx.

Needed because three carriers (Allstate, Travelers, Liberty Mutual) timed out
mid-run on the combined --all-companies search and were retried in fresh
processes that produced per-carrier xlsx files.
"""
from pathlib import Path
from openpyxl import load_workbook, Workbook

OUTPUT_DIR = Path("output")

MAIN = OUTPUT_DIR / "or_all_companies_search.xlsx"
RETRY_FILES = [
    OUTPUT_DIR / "or_allstate_search.xlsx",
    OUTPUT_DIR / "or_travelers_search.xlsx",
    OUTPUT_DIR / "or_liberty_mutual_search.xlsx",
]


def _read_rows(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    header, rows = _read_rows(MAIN)
    print(f"[main] {MAIN.name}: {len(rows)} rows")

    # Drop any rows from failed carriers (they would have been written as 0-row
    # groups but just in case).
    idx_target = header.index("target_company")
    retry_carriers = {"Allstate", "Travelers", "Liberty Mutual"}
    before = len(rows)
    rows = [r for r in rows if r[idx_target] not in retry_carriers]
    print(f"[main] dropped {before - len(rows)} pre-existing rows for retry carriers")

    for path in RETRY_FILES:
        h2, r2 = _read_rows(path)
        if h2 != header:
            raise SystemExit(f"header mismatch in {path}")
        print(f"[merge] {path.name}: {len(r2)} rows")
        rows.extend(r2)

    # Sort by (target_company, serff_tracking_number) for stability.
    idx_serff = header.index("serff_tracking_number")
    rows.sort(key=lambda r: (r[idx_target] or "", r[idx_serff] or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))

    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    # Per-carrier summary
    from collections import Counter
    per = Counter(r[idx_target] for r in rows)
    for c in sorted(per):
        print(f"  {c:16s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
