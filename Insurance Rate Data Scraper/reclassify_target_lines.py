"""Reclassify in_target_lines using strict NAIC TOI/Sub-TOI code matching.

Reads all_states_final.xlsx, applies exact-code matching against the user's
target lists, compares to current in_target_lines, surfaces:
  - Drops (currently True, strict says False)
  - Adds (currently False, strict says True)
  - Edge cases needing user decision (missing TOI, 30.xxxx codes, etc.)
  - Effect on the 23 rows in rate_changes.xlsx

Writes output/strict_classification.json with full per-filing decisions.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import OUTPUT_DIR

SOURCE = OUTPUT_DIR / "all_states_final.xlsx"
RATE_CHANGES = OUTPUT_DIR / "rate_changes.xlsx"
OUT_JSON = OUTPUT_DIR / "strict_classification.json"

# NAIC Uniform Property & Casualty Product Coding Matrix — target sub-TOIs
TARGET_SUB_TOIS = {
    # Personal Auto
    "19.0000",  # Personal Auto Combinations
    "19.0001",  # Private Passenger Auto (PPA)
    "19.0002",  # Motorcycle
    "19.0003",  # Recreational Vehicle (RV)
    "19.0004",  # Other Personal Auto
    # Homeowners
    "04.0000",  # Homeowners Sub-TOI Combinations
    "04.0001",  # Condominium Homeowners
    "04.0002",  # Mobile Homeowners
    "04.0003",  # Owner Occupied Homeowners
    "04.0004",  # Tenant Homeowners
    "04.0005",  # Other Homeowners
}

# Edge-case codes that need explicit user decision before include/exclude
EDGE_SUB_TOIS = {
    "30.0000",  # Homeowner/Auto Combinations (special package)
    "30.1000",  # Dwelling Fire/Personal Liability
}

# Sub-TOIs that look "homeowner-ish" or "auto-ish" by keyword but are explicitly
# different lines per the matrix.
KNOWN_EXCLUDED = {
    "03.0000",  # Personal Farmowners
    "17.0021",  # Personal Umbrella and Excess
    "17.1021",  # Personal Umbrella and Excess
    "17.2021",  # Personal Umbrella and Excess
    "33.0001",  # Other Personal Lines
}


_CODE_RE = re.compile(r"^\s*(\d{2}\.\d{4})\b")


def extract_sub_toi_code(s: Optional[str]) -> Optional[str]:
    """Pull the leading 'NN.NNNN' code off a sub_type_of_insurance string."""
    if not s:
        return None
    m = _CODE_RE.match(str(s))
    return m.group(1) if m else None


_TOI_RE = re.compile(r"^\s*(\d{2}\.\d)\b")


def extract_toi_code(s: Optional[str]) -> Optional[str]:
    """Pull the leading 'NN.N' code off a type_of_insurance string."""
    if not s:
        return None
    m = _TOI_RE.match(str(s))
    return m.group(1) if m else None


def classify(toi_str: Optional[str], sub_str: Optional[str]) -> tuple[str, str]:
    """Return (decision, reason). decision ∈ {include, exclude, edge_case}."""
    sub_code = extract_sub_toi_code(sub_str)
    toi_code = extract_toi_code(toi_str)

    if sub_code in EDGE_SUB_TOIS:
        return ("edge_case", f"sub_toi {sub_code} requires user decision")
    if sub_code in TARGET_SUB_TOIS:
        return ("include", f"sub_toi {sub_code} in target list")
    if sub_code in KNOWN_EXCLUDED:
        return ("exclude", f"sub_toi {sub_code} explicitly excluded")
    if sub_code is None and toi_code is None:
        return ("edge_case", "missing TOI and Sub-TOI codes")
    if sub_code is None:
        return ("edge_case", f"missing Sub-TOI (TOI={toi_code})")
    return ("exclude", f"sub_toi {sub_code} not in target list")


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "1", "yes")


def main() -> int:
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    records = [dict(zip(header, r)) for r in rows[1:]]
    print(f"[load] {SOURCE.name}: {len(records)} filings\n")

    decisions = []
    for r in records:
        toi = r.get("type_of_insurance")
        sub = r.get("sub_type_of_insurance")
        d, reason = classify(toi, sub)
        decisions.append({
            "serff": r.get("serff_tracking_number"),
            "state": r.get("state"),
            "carrier": r.get("target_company"),
            "company_name": r.get("company_name"),
            "toi": toi,
            "sub_toi": sub,
            "toi_code": extract_toi_code(toi),
            "sub_toi_code": extract_sub_toi_code(sub),
            "current_in_target": _to_bool(r.get("in_target_lines")),
            "strict_decision": d,
            "reason": reason,
        })

    # Tallies
    strict_include = [d for d in decisions if d["strict_decision"] == "include"]
    strict_exclude = [d for d in decisions if d["strict_decision"] == "exclude"]
    edge = [d for d in decisions if d["strict_decision"] == "edge_case"]
    cur_true = [d for d in decisions if d["current_in_target"]]

    print("=" * 70)
    print("STRICT CLASSIFICATION SUMMARY")
    print("=" * 70)
    print(f"Strict include:   {len(strict_include)}")
    print(f"Strict exclude:   {len(strict_exclude)}")
    print(f"Edge cases:       {len(edge)}  (need user decision)")
    print(f"Current True:     {len(cur_true)}")

    # Drops: currently True, strict says exclude
    drops = [d for d in decisions if d["current_in_target"] and d["strict_decision"] == "exclude"]
    # Adds: currently False, strict says include
    adds = [d for d in decisions if not d["current_in_target"] and d["strict_decision"] == "include"]
    # Edge cases that are currently True
    edge_in_current = [d for d in edge if d["current_in_target"]]
    edge_not_in_current = [d for d in edge if not d["current_in_target"]]

    print(f"\nDelta vs current in_target_lines:")
    print(f"  Drops  (currently True → strict False): {len(drops)}")
    print(f"  Adds   (currently False → strict True): {len(adds)}")
    print(f"  Edge cases currently True:              {len(edge_in_current)}")
    print(f"  Edge cases currently False:             {len(edge_not_in_current)}")

    if drops:
        print(f"\n=== DROPS ({len(drops)}) — currently in target, strict excludes ===")
        for d in drops[:30]:
            print(f"  {d['serff']:24s} {d['state']:3s} {d['carrier']:14s} sub_toi={d['sub_toi']!r}")
        if len(drops) > 30:
            print(f"  ... and {len(drops)-30} more")

    if adds:
        print(f"\n=== ADDS ({len(adds)}) — currently NOT in target, strict includes ===")
        for d in adds[:30]:
            print(f"  {d['serff']:24s} {d['state']:3s} {d['carrier']:14s} sub_toi={d['sub_toi']!r}")
        if len(adds) > 30:
            print(f"  ... and {len(adds)-30} more")

    if edge:
        print(f"\n=== EDGE CASES ({len(edge)}) — need user decision ===")
        edge_by_reason = defaultdict(list)
        for d in edge:
            edge_by_reason[d["reason"]].append(d)
        for reason, items in edge_by_reason.items():
            print(f"\n  {reason} ({len(items)} filings):")
            for d in items[:20]:
                cur = "True" if d["current_in_target"] else "False"
                print(f"    {d['serff']:24s} {d['state']:3s} {d['carrier']:14s} "
                      f"toi={d['toi']!r:20s} sub_toi={d['sub_toi']!r}  current_target={cur}")
            if len(items) > 20:
                print(f"    ... and {len(items)-20} more")

    # ---- Effect on the 23 rate_changes.xlsx rows ----
    print("\n" + "=" * 70)
    print("EFFECT ON CURRENT 23 ROWS IN rate_changes.xlsx")
    print("=" * 70)
    if RATE_CHANGES.exists():
        wb2 = openpyxl.load_workbook(RATE_CHANGES, read_only=True, data_only=True)
        ws2 = wb2["Rate Changes"]
        rows2 = list(ws2.iter_rows(values_only=True))
        header2 = list(rows2[0])
        rate_rows = [dict(zip(header2, r)) for r in rows2[1:]]
        rate_serffs = {r.get("serff_tracking_number") for r in rate_rows}
        decisions_by_serff = {d["serff"]: d for d in decisions}
        in_strict = []
        out_strict = []
        edge_in_rate = []
        for serff in rate_serffs:
            d = decisions_by_serff.get(serff)
            if not d:
                print(f"  ! {serff} not found in source workbook")
                continue
            if d["strict_decision"] == "include":
                in_strict.append(d)
            elif d["strict_decision"] == "edge_case":
                edge_in_rate.append(d)
            else:
                out_strict.append(d)
        print(f"  Of 23 current rows:")
        print(f"    Pass strict:   {len(in_strict)}")
        print(f"    Fail strict:   {len(out_strict)}")
        print(f"    Edge case:     {len(edge_in_rate)}")
        if out_strict:
            print(f"\n  ROWS THAT WOULD DROP:")
            for d in out_strict:
                print(f"    {d['serff']:24s} sub_toi={d['sub_toi']!r}  reason={d['reason']}")
        if edge_in_rate:
            print(f"\n  EDGE-CASE ROWS:")
            for d in edge_in_rate:
                print(f"    {d['serff']:24s} sub_toi={d['sub_toi']!r}  reason={d['reason']}")

    # ---- Sub-TOI breakdown of strict-include set ----
    print("\n=== STRICT INCLUDE BREAKDOWN BY SUB-TOI ===")
    sub_counts = Counter(d["sub_toi_code"] for d in strict_include)
    for code, n in sorted(sub_counts.items()):
        # Find a representative full label
        label_match = next(
            (d["sub_toi"] for d in strict_include if d["sub_toi_code"] == code),
            code,
        )
        print(f"  {n:4d}  {label_match}")

    OUT_JSON.write_text(json.dumps(decisions, indent=2, default=str), encoding="utf-8")
    print(f"\n[write] {OUT_JSON}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
