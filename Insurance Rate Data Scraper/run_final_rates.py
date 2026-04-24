"""Build {state}_final_rates.xlsx from cached/downloaded SERFF system PDFs.

Usage:
    python run_final_rates.py ID
    python run_final_rates.py WA
    python run_final_rates.py CO

Reads the state's intermediate file `output/{state.lower()}_final.xlsx`,
filters to target-TOI target-carrier filings, downloads each filing's
system Filing Summary PDF (cached), parses it via
`utils.parse_filing_summary_pdf`, and emits one row per per-company rate
row in AM Best Disposition Page Data format.

Non-rate filings (Form, Rule, new-product Rate/Rule) are excluded.
"""
from __future__ import annotations

import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
import pdfplumber
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import HEADLESS, USER_AGENT
from src.detail import download_system_summary_pdf
from src.search import (
    _back_to_results,
    _click_row_to_detail,
    _set_rows_per_page_100,
    _submit_search,
)
from src.utils import parse_filing_summary_pdf

TARGET_TOI = ("19.0", "04.0")  # Personal Auto + Homeowners (Farmowners explicitly out of scope)
GROUP_SEARCH = {  # group -> list of SERFF search terms (each term = a separate SERFF query)
    "State Farm":     ["state farm"],
    "GEICO":          ["geico"],
    # Encompass files under its own brand on SERFF and is NOT returned by an
    # "allstate" keyword search; we search both names under the Allstate group.
    "Allstate":       ["allstate", "encompass"],
    "Travelers":      ["travelers"],
    # Safeco is Liberty Mutual's independent-agent brand and files under its
    # own name; it does NOT surface under a "liberty mutual" search.
    "Liberty Mutual": ["liberty mutual", "safeco"],
    "Progressive":    ["progressive"],
}
GROUP_KW = {  # subsidiary-name keywords used to assign a filing to its parent group
    "State Farm":     ["state farm", "mga insurance"],
    "GEICO":          ["geico"],
    "Allstate":       ["allstate", "encompass", "integon", "north american insurance"],
    "Travelers":      ["travelers", "standard fire"],
    "Liberty Mutual": ["liberty mutual", "safeco", "first national insurance company of america", "general insurance company of america", "american states", "american economy"],
    "Progressive":    ["progressive"],
}
# Out-of-scope subsidiaries (do NOT classify as one of our groups):
#   - Esurance (Allstate, wound down 2020)
#   - Drive Insurance (Progressive, retired)
#   - United Financial (Progressive specialty)
# Identifies filings that launch a new product (vs. modifying an existing one).
# A bare "Introduction of" / "Initial Submission" / "Initial Filing" keyword can
# false-positive on body text describing rating-factor additions, deductible
# tweaks, or references to prior filings. This regex anchors those keywords to
# header fields (Project Name/Number, Company Tracking #) and requires body-text
# "introduction of" to be followed by a product-launch noun (Program, line of
# business). Standalone "New Program" / "new product" remain catch-alls because
# audit found no false-positives for those phrasings in our corpus.
NEW_PRODUCT_RE = re.compile(
    r"("
    r"Project Name/Number:[^\n]*\b(?:Initial Filing|Initial Submission|Introduction of)\b"
    r"|Company Tracking #:[^\n]*\bINTRODUCTION OF\b"
    r"|\bNew Program\b"
    r"|\bnew product\b"
    r"|\bintroduction of\b[\s\S]{0,120}\b(?:Program|line of business|lines of business)\b"
    r")",
    re.IGNORECASE,
)
RATE_FILING_TYPES = {"Rate", "Rate/Rule"}
PDF_FILING_TYPE_RE = re.compile(r"Filing Type:\s*([A-Za-z/ \-]+)\s*$", re.MULTILINE)


def carrier_group(*names: Optional[str]) -> Optional[str]:
    """Match the first non-empty name against carrier-group keywords.
    Multi-company filings list company_name as 'Multiple' — pass target_company
    as a fallback so they're not dropped."""
    for name in names:
        n = (name or "").lower()
        if not n:
            continue
        for g, kws in GROUP_KW.items():
            if any(k in n for k in kws):
                return g
    return None


@dataclass
class Target:
    tracking: str
    filing_id: str
    company: str
    toi: str
    sub_toi: str
    filing_type_xlsx: str
    submission_date: object
    disposition_date: object
    disposition_status_xlsx: str
    group: str


def load_targets(state: str) -> list[Target]:
    src = Path(f"output/{state.lower()}_final.xlsx")
    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        toi = r[ix["type_of_insurance"]] or ""
        if not any(toi.startswith(p) for p in TARGET_TOI):
            grp = carrier_group(r[ix["company_name"]], r[ix["target_company"]])
            if grp and toi:
                tk = r[ix["serff_tracking_number"]] or ""
                print(f"  excluded — out of scope TOI: {tk} ({grp}, {toi})", flush=True)
            continue
        grp = carrier_group(r[ix["company_name"]], r[ix["target_company"]])
        if not grp:
            continue
        out.append(Target(
            tracking=r[ix["serff_tracking_number"]] or "",
            filing_id=str(r[ix["filing_id"]] or ""),
            company=r[ix["company_name"]] or "",
            toi=toi,
            sub_toi=r[ix["sub_type_of_insurance"]] or "",
            filing_type_xlsx=r[ix["filing_type"]] or "",
            submission_date=r[ix["submission_date"]],
            disposition_date=r[ix["disposition_date"]],
            disposition_status_xlsx=r[ix["disposition_status"]] or "",
            group=grp,
        ))
    return out


def download_all_pdfs(state: str, targets: list[Target]) -> dict[str, str]:
    """Download system PDF for every target. Returns {filing_id: download_status}."""
    by_group: dict[str, list[Target]] = {}
    for t in targets:
        by_group.setdefault(t.group, []).append(t)
    statuses: dict[str, str] = {}
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
        page = ctx.new_page()

        def _refresh_context():
            nonlocal ctx, page
            try:
                ctx.close()
            except Exception:
                pass
            ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
            page = ctx.new_page()

        def _submit_with_retry(st: str, term: str) -> bool:
            for attempt in range(3):
                try:
                    if _submit_search(page, st, term):
                        return True
                except Exception as e:
                    print(f"    [retry {attempt+1}/3] submit_search {term!r}: {e}", flush=True)
                _refresh_context()
            return False

        for grp, items in by_group.items():
            uncached = []
            for t in items:
                pdf = Path(f"output/pdfs/{state}/{t.filing_id}/filing_summary.pdf")
                if pdf.exists() and pdf.stat().st_size > 5000:
                    statuses[t.filing_id] = "cached"
                else:
                    uncached.append(t)
            if not uncached:
                print(f"[{grp}] all {len(items)} cached", flush=True); continue
            search_terms = GROUP_SEARCH[grp]
            print(f"[{grp}] searches={search_terms!r}, downloading {len(uncached)}/{len(items)}", flush=True)
            remaining = list(uncached)
            for search_term in search_terms:
                if not remaining:
                    break
                print(f"  [{grp}] search={search_term!r}, attempting {len(remaining)} filing(s)", flush=True)
                if not _submit_with_retry(state, search_term):
                    print(f"  [{grp}] search submission failed for {search_term!r} after retries", flush=True)
                    continue
                _set_rows_per_page_100(page)
                still_remaining: list[Target] = []
                for idx, t in enumerate(remaining, 1):
                    found = False
                    for _ in range(10):
                        if page.locator(f'tr[data-rk="{t.filing_id}"]').count():
                            found = True; break
                        nxt = page.locator(".ui-paginator-next").first
                        if not nxt.count() or "ui-state-disabled" in (nxt.get_attribute("class") or ""):
                            break
                        nxt.click(); page.wait_for_load_state("networkidle", timeout=15000)
                    if not found:
                        still_remaining.append(t)
                        continue
                    dest_dir = Path(f"output/pdfs/{state}/{t.filing_id}")
                    pdf = download_system_summary_pdf(page, t.filing_id, t.tracking, dest_dir)
                    statuses[t.filing_id] = "ok" if pdf else "fail:download"
                    print(f"    [{idx}/{len(remaining)}] {t.tracking}: {statuses[t.filing_id]}", flush=True)
                    if not page.locator(".ui-paginator-next").count():
                        _submit_with_retry(state, search_term); _set_rows_per_page_100(page)
                remaining = still_remaining
            for t in remaining:
                statuses[t.filing_id] = "fail:row_not_found"
                print(f"  {t.tracking}: not found in any of {search_terms!r}", flush=True)
        browser.close()
    return statuses


def detect_filing_type_and_new_product(pdf_path: Path) -> tuple[Optional[str], bool]:
    """Read PDF text once, return (filing_type, is_new_product)."""
    with pdfplumber.open(str(pdf_path)) as pdf:
        text = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    ft = None
    if m := PDF_FILING_TYPE_RE.search(text):
        ft = m.group(1).strip()
    return ft, bool(NEW_PRODUCT_RE.search(text))


# AM Best Disposition Page Data column order (per user spec)
COLUMNS = [
    "state",
    "effective_date",
    "company_name",
    "line_of_business",
    "sub_type_of_insurance",
    "overall_indicated_change",
    "overall_rate_impact",
    "written_premium_change",
    "policyholders_affected",
    "written_premium_for_program",
    "maximum_percent_change",
    "minimum_percent_change",
    "rate_activity",
    "serff_tracking_number",
    "disposition_status",
    "filing_date",
    "source_pdf",
]


def build_rows(state: str, targets: list[Target]) -> tuple[list[dict], dict]:
    """Parse each cached PDF and emit one row per per-company rate row.
    Returns (rows, stats)."""
    rows: list[dict] = []
    stats = {
        "filings_total": len(targets),
        "filings_excluded_form_or_rule": 0,
        "filings_excluded_new_product": 0,
        "filings_excluded_no_pdf": 0,
        "filings_excluded_rate_data_does_not_apply": 0,
        "filings_emitted": 0,
        "rows_emitted": 0,
        "anchor_match_count": 0,
    }
    for t in targets:
        pdf = Path(f"output/pdfs/{state}/{t.filing_id}/filing_summary.pdf")
        if not pdf.exists() or pdf.stat().st_size < 5000:
            stats["filings_excluded_no_pdf"] += 1; continue
        ft_pdf, is_new = detect_filing_type_and_new_product(pdf)
        ft = ft_pdf or t.filing_type_xlsx
        if ft not in RATE_FILING_TYPES:
            stats["filings_excluded_form_or_rule"] += 1; continue
        if is_new:
            stats["filings_excluded_new_product"] += 1; continue
        fs = parse_filing_summary_pdf(pdf, t.tracking)
        if not fs.rate_data_applies:
            stats["filings_excluded_rate_data_does_not_apply"] += 1; continue
        if not fs.company_rates:
            # rate_data_applies=True but no rows extracted — record as zero-row anomaly
            print(f"  ! {t.tracking}: rate_data_applies=True but 0 rows extracted")
            stats["filings_excluded_no_pdf"] += 1
            continue

        # Determine rate_activity from disposition status
        ds = (fs.disposition_status or "").upper()
        if "WITHDRAWN" in ds:
            activity = "rate_change_withdrawn"
        elif "DISAPPROV" in ds:
            activity = "rate_change_disapproved"
        elif "PENDING" in ds:
            activity = "rate_change_pending"
        else:
            activity = "rate_change"

        eff = fs.effective_date_new or fs.effective_date_renewal
        rel_pdf = pdf.relative_to(Path(".")).as_posix() if pdf.is_absolute() is False else pdf.as_posix()
        for r in fs.company_rates:
            rows.append({
                "state": state,
                "effective_date": eff,
                "company_name": r.company_name,
                "line_of_business": t.toi,
                "sub_type_of_insurance": t.sub_toi,
                "overall_indicated_change": r.overall_indicated_change,
                "overall_rate_impact": r.overall_rate_impact,
                "written_premium_change": r.written_premium_change,
                "policyholders_affected": r.policyholders_affected,
                "written_premium_for_program": r.written_premium_for_program,
                "maximum_percent_change": r.maximum_pct_change,
                "minimum_percent_change": r.minimum_pct_change,
                "rate_activity": activity,
                "serff_tracking_number": t.tracking,
                "disposition_status": fs.disposition_status,
                "filing_date": (t.submission_date.isoformat() if hasattr(t.submission_date, "isoformat") else t.submission_date),
                "source_pdf": rel_pdf,
            })
        stats["filings_emitted"] += 1
        stats["rows_emitted"] += len(fs.company_rates)
    return rows, stats


def write_xlsx(rows: list[dict], state: str) -> Path:
    out = Path(f"output/{state.lower()}_final_rates.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "rates"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    wb.save(out)
    return out


# ============================================================
# Idaho-specific anchor verification (SFMA-134676753 vs AM Best)
# ============================================================
ANCHOR_EXPECTED = {
    ("SFMA-134676753", "State Farm Fire and Casualty Company"): dict(
        ind="15.900%", imp="-2.100%", prem_chg="-554469", ph=20679,
        prem_for="26357498", maxp="388.400%", minp="-41.500%",
    ),
    ("SFMA-134676753", "State Farm Mutual Automobile Insurance Company"): dict(
        ind="-2.600%", imp="-9.700%", prem_chg="-25716996", ph=360274,
        prem_for="263832752", maxp="847.900%", minp="-52.200%",
    ),
}


def verify_anchor(rows: list[dict]) -> tuple[int, list[str]]:
    """Return (matches_count, mismatch_messages) for anchor SFMA-134676753."""
    mismatches: list[str] = []
    matched = 0
    by_key = {(r["serff_tracking_number"], r["company_name"]): r for r in rows}
    for key, exp in ANCHOR_EXPECTED.items():
        r = by_key.get(key)
        if not r:
            mismatches.append(f"  MISSING: {key}"); continue
        actual = dict(
            ind=r["overall_indicated_change"], imp=r["overall_rate_impact"],
            prem_chg=r["written_premium_change"], ph=r["policyholders_affected"],
            prem_for=r["written_premium_for_program"],
            maxp=r["maximum_percent_change"], minp=r["minimum_percent_change"],
        )
        ok = True
        for k, ev in exp.items():
            if actual[k] != ev:
                mismatches.append(f"  {key} {k}: got={actual[k]!r} expected={ev!r}"); ok = False
        if ok: matched += 7  # 7 fields per row
    return matched, mismatches


def main():
    state = (sys.argv[1] if len(sys.argv) > 1 else "ID").upper()
    t0 = time.time()
    print(f"=== {state} final-rates pipeline ===")
    targets = load_targets(state)
    print(f"loaded {len(targets)} target-TOI target-carrier filings")
    download_all_pdfs(state, targets)
    rows, stats = build_rows(state, targets)
    out = write_xlsx(rows, state)
    elapsed = time.time() - t0

    # storage delta — sum size of cached system PDFs
    pdf_dir = Path(f"output/pdfs/{state}")
    total_kb = sum(p.stat().st_size for p in pdf_dir.rglob("filing_summary.pdf")) / 1024

    # field-completion rates
    completion = {c: 0 for c in COLUMNS}
    for r in rows:
        for c in COLUMNS:
            if r.get(c) not in (None, ""):
                completion[c] += 1

    print("\n=== STATS ===")
    for k, v in stats.items(): print(f"  {k}: {v}")
    print(f"\n=== FIELD COMPLETION ({len(rows)} rows) ===")
    for c in COLUMNS:
        pct = (100 * completion[c] / len(rows)) if rows else 0
        print(f"  {completion[c]:4d}/{len(rows)}  ({pct:5.1f}%)  {c}")
    print(f"\n=== STORAGE ===")
    print(f"  system PDFs cached: {total_kb:.1f} KB across {len(targets)} filings")
    print(f"  avg per filing:     {total_kb / max(1, len(targets)):.1f} KB")
    print(f"\n=== RUNTIME ===")
    print(f"  elapsed: {elapsed:.1f} s ({elapsed/60:.1f} min)")
    print(f"\n=== OUTPUT ===")
    print(f"  -> {out} ({len(rows)} data rows)")

    if state == "ID":
        matched, mismatches = verify_anchor(rows)
        print(f"\n=== ANCHOR (SFMA-134676753 vs AM Best) ===")
        print(f"  matched: {matched}/14 fields")
        if mismatches:
            print("  mismatches:")
            for m in mismatches: print(m)


if __name__ == "__main__":
    main()
