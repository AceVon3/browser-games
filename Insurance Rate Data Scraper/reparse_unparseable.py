"""Re-run PDF parser against already-downloaded WA filings, no re-scrape.

Operates on every Rate/Rule filing in `output/wa_final.xlsx` (both the Filings
sheet and the Unparseable PDFs sheet) so we can detect regressions on filings
that previously parsed cleanly.

Per-PDF priority tiers:
  * MEMO  — names containing memo/summary/cover letter/justification/filing packet
  * SKIP  — names containing manual/tracked changes/rate pages/exhibit/complete/compare
  * DEFAULT — everything else
If any MEMO PDF exists for a filing, only MEMO + DEFAULT are parsed (SKIP is
dropped entirely). If no MEMO exists, all three tiers are tried in order.

Each PDF is extracted in a subprocess with a 60-second hard timeout (see
`extract_pdf_text_with_timeout`) so pdfplumber hangs are bounded.

Reports:
  * how many of the 26 unparseable filings became parsed / premium-neutral
  * how many remain unparseable / timed-out / had no usable PDFs
  * delta against previously-parsed filings — flagging any regression where
    a previously-set overall_rate_effect went missing or changed value
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

from src.utils import parse_rate_effect_pdf

LARGE_PDF_SKIP_PARSE_MB = 15.0
PDF_DIR = Path("output/pdfs/WA")
EXCEL = Path("output/wa_final.xlsx")
PER_PDF_TIMEOUT_S = 60.0

MEMO_KEYWORDS = (
    "memo", "summary", "cover letter", "justification", "filing packet",
)
SKIP_KEYWORDS = (
    "manual", "tracked changes", "rate pages", "exhibit", "complete", "compare",
)


def categorize(name: str) -> str:
    n = name.lower()
    if any(k in n for k in MEMO_KEYWORDS):
        return "memo"
    if any(k in n for k in SKIP_KEYWORDS):
        return "skip"
    return "default"


def prioritize(paths: list[Path]) -> list[Path]:
    memos, defaults, skips = [], [], []
    for p in paths:
        c = categorize(p.name)
        (memos if c == "memo" else defaults if c == "default" else skips).append(p)
    memos.sort(key=lambda p: p.name)
    defaults.sort(key=lambda p: p.name)
    skips.sort(key=lambda p: p.name)
    if memos:
        return memos + defaults
    return defaults + skips


def parse_filing_pdfs(filing_id: str, verbose: bool = True) -> tuple[str, dict, list[str]]:
    """Returns (new_status, parsed_fields, log_lines)."""
    dest_dir = PDF_DIR / filing_id
    log: list[str] = []
    if not dest_dir.exists():
        return "no_pdfs_attached", {}, [f"  [no dir] {dest_dir}"]
    available = list(dest_dir.glob("*.pdf"))
    if not available:
        return "no_pdfs_attached", {}, ["  [no pdfs in dir]"]
    ordered = prioritize(available)

    fields: dict = {}
    parse_attempted = 0
    timeouts = 0
    for pdf in ordered:
        size_mb = pdf.stat().st_size / (1024 * 1024)
        if size_mb > LARGE_PDF_SKIP_PARSE_MB:
            if verbose:
                log.append(f"  [skip {size_mb:.1f}MB oversize] {pdf.name}")
            continue
        parse_attempted += 1
        try:
            result, status = parse_rate_effect_pdf(
                pdf, filing_id, timeout_s=PER_PDF_TIMEOUT_S,
            )
        except Exception as e:
            log.append(f"  [exception] {pdf.name}: {type(e).__name__} {e}")
            continue
        if status == "timeout":
            timeouts += 1
            log.append(f"  [TIMEOUT] {pdf.name}")
            continue
        new_keys = [k for k in result if k not in fields]
        if new_keys and verbose:
            log.append(
                f"  [hit {pdf.name}] " + ", ".join(f"{k}={result[k]}" for k in new_keys)
            )
        for k, v in result.items():
            fields.setdefault(k, v)

    if fields:
        return "parsed", fields, log
    if parse_attempted == 0:
        return "all_pdfs_too_large_to_parse", fields, log
    if timeouts > 0 and timeouts == parse_attempted:
        return "timeout_skipped", fields, log
    if timeouts > 0:
        return "no_fields_matched_with_timeouts", fields, log
    return "no_fields_matched", fields, log


def load_filings_index() -> dict[str, dict]:
    """Returns {filing_id: {serff, status, ore, requested, approved, ...}}."""
    wb = openpyxl.load_workbook(EXCEL, read_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    out: dict[str, dict] = {}
    for r in rows[1:]:
        fid = r[idx["filing_id"]]
        if not fid:
            continue
        out[str(fid)] = {
            "serff": r[idx["serff_tracking_number"]],
            "company": r[idx["company_name"]],
            "filing_type": r[idx["filing_type"]],
            "status": r[idx["pdf_parse_status"]],
            "ore": r[idx.get("overall_rate_effect", -1)] if "overall_rate_effect" in idx else None,
            "req": r[idx.get("requested_rate_effect", -1)] if "requested_rate_effect" in idx else None,
            "appr": r[idx.get("approved_rate_effect", -1)] if "approved_rate_effect" in idx else None,
        }
    return out


def load_unparseable_ids() -> set[str]:
    wb = openpyxl.load_workbook(EXCEL, read_only=True)
    ws = wb["Unparseable PDFs"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    return {str(r[idx["filing_id"]]) for r in rows[1:]}


def fmt_val(v):
    if v is None:
        return "None"
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def main():
    filings = load_filings_index()
    unparseable_ids = load_unparseable_ids()
    print(f"Loaded {len(filings)} total filings; {len(unparseable_ids)} in unparseable queue")

    queue_results: dict[str, dict] = {}
    regression_results: dict[str, dict] = {}

    # Re-run on the unparseable queue with verbose logging
    print("\n" + "=" * 60)
    print(f"PHASE 1 — re-parse {len(unparseable_ids)} unparseable filings")
    print("=" * 60)
    for fid in sorted(unparseable_ids):
        info = filings.get(fid, {})
        serff = info.get("serff", "?")
        print(f"\n--- {serff}  (filing_id={fid}) ---")
        status, fields, log = parse_filing_pdfs(fid, verbose=True)
        for line in log:
            print(line)
        print(f"  -> status={status}, fields={fields}")
        queue_results[fid] = {"status": status, "fields": fields, "info": info}

    # Re-run on all OTHER Rate/Rule filings to detect regressions
    rate_rule_ids = [
        fid for fid, info in filings.items()
        if info.get("filing_type") == "Rate/Rule" and fid not in unparseable_ids
    ]
    print("\n" + "=" * 60)
    print(f"PHASE 2 — regression check on {len(rate_rule_ids)} previously-parsed filings")
    print("=" * 60)
    regressions: list[str] = []
    for i, fid in enumerate(sorted(rate_rule_ids), 1):
        info = filings[fid]
        serff = info.get("serff", "?")
        status, fields, log = parse_filing_pdfs(fid, verbose=False)
        regression_results[fid] = {"status": status, "fields": fields, "info": info}
        prev_ore = info.get("ore")
        new_ore = fields.get("overall_rate_effect")
        # Regression = previously had ore set, now missing OR different value
        if prev_ore is not None and prev_ore != "":
            try:
                prev_f = float(prev_ore)
            except (TypeError, ValueError):
                prev_f = None
            if prev_f is not None:
                if new_ore is None:
                    regressions.append(
                        f"  REGRESSION {serff}: previously overall_rate_effect={prev_f}, "
                        f"now MISSING (new status={status})"
                    )
                elif abs(prev_f - float(new_ore)) > 0.01:
                    regressions.append(
                        f"  REGRESSION {serff}: previously overall_rate_effect={prev_f}, "
                        f"now {new_ore}"
                    )
        if i % 20 == 0:
            print(f"  ... {i}/{len(rate_rule_ids)} regression-checked")

    # ---------- summary ----------
    print("\n" + "=" * 60)
    print("UNPARSEABLE QUEUE — RE-PARSE SUMMARY")
    print("=" * 60)
    n_total = len(unparseable_ids)
    status_counts: Counter = Counter(r["status"] for r in queue_results.values())
    now_parsed_nonzero = []
    now_parsed_zero = []
    still_unparseable = []
    timed_out = []
    other = []
    for fid, r in queue_results.items():
        info = r["info"]
        serff = info.get("serff", fid)
        ore = r["fields"].get("overall_rate_effect")
        if r["status"] == "parsed":
            if ore == 0.0:
                now_parsed_zero.append((serff, r["fields"]))
            else:
                now_parsed_nonzero.append((serff, r["fields"]))
        elif r["status"] in ("no_fields_matched", "no_fields_matched_with_timeouts"):
            still_unparseable.append((serff, r["status"]))
        elif r["status"] in ("timeout_skipped",):
            timed_out.append((serff, r["status"]))
        else:
            other.append((serff, r["status"]))

    print(f"\nQueue size:                 {n_total}")
    pct = lambda n: f"({100*n/n_total:.1f}%)" if n_total else ""
    print(f"Now parsed (non-zero):      {len(now_parsed_nonzero)} {pct(len(now_parsed_nonzero))}")
    print(f"Now premium-neutral (0%):   {len(now_parsed_zero)} {pct(len(now_parsed_zero))}")
    print(f"Still unparseable:          {len(still_unparseable)} {pct(len(still_unparseable))}")
    print(f"Timed out (all PDFs):       {len(timed_out)} {pct(len(timed_out))}")
    print(f"Other (no PDFs/oversize):   {len(other)} {pct(len(other))}")

    print(f"\nStatus distribution: {dict(status_counts)}")

    print("\n--- Newly parsed (non-zero overall_rate_effect) ---")
    for serff, fields in now_parsed_nonzero:
        ore = fields.get("overall_rate_effect")
        extra = {k: v for k, v in fields.items() if k != "overall_rate_effect"}
        print(f"  {serff}  ore={ore}  other={extra or '{}'}")

    print("\n--- Newly premium-neutral (0.0%) ---")
    for serff, fields in now_parsed_zero:
        print(f"  {serff}  fields={fields}")

    print("\n--- Still unparseable ---")
    for serff, status in still_unparseable:
        print(f"  {serff}  ({status})")

    if timed_out:
        print("\n--- Timed out ---")
        for serff, status in timed_out:
            print(f"  {serff}  ({status})")

    print("\n" + "=" * 60)
    print("REGRESSION CHECK (previously-parsed filings)")
    print("=" * 60)
    print(f"Filings checked:     {len(regression_results)}")
    print(f"Regressions found:   {len(regressions)}")
    if regressions:
        for line in regressions:
            print(line)
    else:
        print("  None — no previously-parsed filing changed value or lost it.")


if __name__ == "__main__":
    main()
