"""Audit the 23 rate_changes.xlsx filings: classify each value as
overall_impact / base_rate / indicated / ambiguous by re-reading PDF memos.

Walks output/pdfs/{state}/{filing_id}/*.pdf in memo-priority order, scans for
labeled rate-change phrases, captures the nearby percentage + a 200-char
context snippet, and writes:
    output/rate_change_audit.json  — full per-filing evidence dump
    stdout                         — summary table + ambiguous samples

Does NOT modify rate_changes.xlsx. The user reviews findings before rebuild.
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import OUTPUT_DIR
from src.utils import extract_pdf_text_with_timeout

SOURCE = OUTPUT_DIR / "rate_changes.xlsx"
ALL_STATES = OUTPUT_DIR / "all_states_final.xlsx"
PDF_ROOT = OUTPUT_DIR / "pdfs"
AUDIT_JSON = OUTPUT_DIR / "rate_change_audit.json"

PER_PDF_TIMEOUT_S = 45.0
LARGE_PDF_SKIP_MB = 15.0

MEMO_KEYWORDS = ("memo", "summary", "cover letter", "justification", "filing packet")
SKIP_KEYWORDS = ("manual", "tracked changes", "rate pages", "exhibit", "complete", "compare")


# ---------- classification patterns ----------
# Each pattern captures (the percentage, any qualifier word). Patterns are
# probed in declared order; we record EVERY match (we want the full evidence
# for classification, not first-match-wins).

_PCT = r"([+\-]?\(?\s*\d+(?:\.\d+)?\s*%\)?)"

# pdfplumber sometimes drops spaces between words in tight-kerned PDFs (e.g.
# "overallrateimpact" instead of "overall rate impact"). Patterns use \s* (zero
# or more) instead of \s+ so they tolerate both spaced and smushed text.
# We also pre-normalize text by inserting spaces at lowercase->Uppercase
# boundaries before scanning.

OVERALL_IMPACT_PATTERNS = [
    # PROPOSED/ADOPTED overall change phrases. Carefully exclude "indication"
    # since "Overall Rate Indication" is the INDICATED metric, not the proposed.
    re.compile(rf"overall\s*rate\s*(?:change|impact|effect|increase|decrease|adjustment)\s*of\s*{_PCT}", re.IGNORECASE),
    re.compile(rf"{_PCT}\s*overall\s*rate\s*(?:change|impact|effect|increase|decrease|adjustment)", re.IGNORECASE),
    # "overall X% change to [Company]" / "overall X% change for the [Program]"
    re.compile(rf"overall\s*{_PCT}\s*(?:rate\s*)?(?:change|increase|decrease|adjustment|impact)\s*(?:to|for)\s+", re.IGNORECASE),
    # "result in an X% overall rate"
    re.compile(rf"result\s*in\s*an?\s*{_PCT}\s*overall\s*rate", re.IGNORECASE),
    # "result in an overall rate change of X%"
    re.compile(rf"result\s*in\s*an?\s*overall\s*rate\s*(?:change|impact|effect|increase|decrease)\s*of\s*{_PCT}", re.IGNORECASE),
    # "result in a statewide average X% change"
    re.compile(rf"result\s*in\s*an?\s*statewide\s*(?:average\s*)?{_PCT}\s*change", re.IGNORECASE),
    # "statewide average X% change for the [Program]"
    re.compile(rf"statewide\s*(?:average\s*)?{_PCT}\s*change\s*(?:for|to)\s+", re.IGNORECASE),
    # "filing includes a revision of X% to the/our premium income level"
    re.compile(rf"filing\s*includes\s*(?:a\s*)?revision\s*of\s*{_PCT}\s*to\s*(?:our|the)?\s*premium\s*income", re.IGNORECASE),
    # "Overall Rate Level Change/Impact" header (NOT "Indication")
    re.compile(rf"overall\s*rate\s*level\s*(?:change|impact|increase|decrease)[^\n%]{{0,40}}?{_PCT}", re.IGNORECASE),
    # "Total Rate Level Change/Impact" (NOT "indication")
    re.compile(rf"total\s*rate\s*(?:level\s*)?(?:change|impact|effect|increase|decrease)\s*of\s*{_PCT}", re.IGNORECASE),
    # "weighted average rate change of X%"
    re.compile(rf"weighted\s*average\s*rate\s*(?:change|impact|effect|increase|decrease)\s*of\s*{_PCT}", re.IGNORECASE),
    # "proposed rate change of X%" / "adopted rate change of X%"
    re.compile(rf"(?:propos|adopt)(?:ed|ing)\s*(?:overall\s*)?rate\s*(?:change|increase|decrease|impact)\s*of\s*{_PCT}", re.IGNORECASE),
    # "Program Rate Change of Changes ... X%" (table-style)
    re.compile(rf"program\s*rate\s*change[^\n%]{{0,40}}?{_PCT}", re.IGNORECASE),
]

BASE_RATE_PATTERNS = [
    re.compile(rf"base\s*rate\s*(?:change|increase|decrease|adjustment|impact)\s*of\s*{_PCT}", re.IGNORECASE),
    re.compile(rf"{_PCT}\s*base\s*rate\s*(?:change|increase|decrease|adjustment)", re.IGNORECASE),
    re.compile(rf"base\s*rate[s]?\s*(?:are|will\s*be)\s*(?:increas|decreas|adjust)\w*\s*(?:by\s*)?{_PCT}", re.IGNORECASE),
    re.compile(rf"increas\w+\s*(?:our\s*|the\s*)?base\s*rates?\s*by\s*{_PCT}", re.IGNORECASE),
    re.compile(rf"baserate(?:change|increase|decrease|adjustment)[^\n%]{{0,40}}?{_PCT}", re.IGNORECASE),
]

INDICATED_PATTERNS = [
    # Standard "indicated rate change of X%" phrases
    re.compile(rf"indicated\s*rate\s*(?:level\s*)?(?:change|increase|decrease|impact|effect)\s*of\s*{_PCT}", re.IGNORECASE),
    # "indicated rate level change ... is X%"  / "indicated rate change for all coverages combined is X%"
    re.compile(rf"indicated\s*rate\s*(?:level\s*)?change[^.]{{0,80}}?\s+is\s+{_PCT}", re.IGNORECASE),
    # "Overall Rate Indication X%" / "Overall Indication X%" — labeled overall but it's the INDICATED metric, not adopted
    re.compile(rf"overall\s*(?:rate\s*)?indication[\s\-:]*{_PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*indication\s*is\s*{_PCT}", re.IGNORECASE),
    # "credibility-weighted indication of X%"
    re.compile(rf"credibility[-\s]*weighted\s*indication\s*(?:of|is)\s*{_PCT}", re.IGNORECASE),
    re.compile(rf"credibility[-\s]*weighted\s*indicated\s*rate\s*(?:change|increase)\s*of\s*{_PCT}", re.IGNORECASE),
    # "actuarially indicated X%"
    re.compile(rf"actuarial(?:ly)?\s*indicated\s*(?:rate\s*)?(?:change|increase|decrease)\s*of\s*{_PCT}", re.IGNORECASE),
    # "rate indication of X%" / "rate indication is X%"
    re.compile(rf"rate\s*indication\s*(?:of|is)\s*{_PCT}", re.IGNORECASE),
    # Smushed: "credibility-weightedindication"
    re.compile(rf"credibility[-\s]*weightedindication[^\n%]{{0,40}}?{_PCT}", re.IGNORECASE),
    # "Indicated Change = X%" (table format)
    re.compile(rf"indicated\s*change\s*=\s*[^\n%]{{0,30}}?{_PCT}", re.IGNORECASE),
    # "current rate level indication ... is X%"
    re.compile(rf"(?:current|most\s*current)\s*rate\s*level\s*indication[^.]{{0,80}}?\s*is\s*{_PCT}", re.IGNORECASE),
]

REQUESTED_PATTERNS = [
    re.compile(rf"requested\s*rate\s*(?:change|increase|decrease|impact|effect)\s*of\s*{_PCT}", re.IGNORECASE),
    re.compile(rf"requesting\s*(?:an?\s*)?(?:overall\s*)?rate\s*(?:change|increase|decrease|adjustment)\s*of\s*{_PCT}", re.IGNORECASE),
    re.compile(rf"propos(?:ing|ed)\s*(?:an?\s*)?rate\s*(?:change|increase|decrease|adjustment)\s*of\s*{_PCT}", re.IGNORECASE),
]

APPROVED_PATTERNS = [
    re.compile(rf"approved\s*rate\s*(?:change|increase|decrease|impact|effect)\s*of\s*{_PCT}", re.IGNORECASE),
]


def _normalize_text(text: str) -> str:
    """Insert spaces at lowercase->Uppercase boundaries and around digits to
    repair pdfplumber's space-dropping on tight-kerned PDFs."""
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = re.sub(r"([a-zA-Z])(\d)", r"\1 \2", text)
    text = re.sub(r"(\d)([a-zA-Z])", r"\1 \2", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _parse_pct(s: str) -> Optional[float]:
    if not s:
        return None
    raw = s.replace(" ", "")
    neg = "(" in raw and ")" in raw
    m = re.search(r"[+\-]?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        v = float(m.group(0))
    except ValueError:
        return None
    if neg and v > 0:
        v = -v
    return v


def _categorize(name: str) -> str:
    n = name.lower()
    if any(k in n for k in MEMO_KEYWORDS):
        return "memo"
    if any(k in n for k in SKIP_KEYWORDS):
        return "skip"
    return "default"


def _prioritize(paths: list[Path]) -> list[Path]:
    memos, defaults, skips = [], [], []
    for p in paths:
        c = _categorize(p.name)
        (memos if c == "memo" else defaults if c == "default" else skips).append(p)
    memos.sort(key=lambda p: p.name)
    defaults.sort(key=lambda p: p.name)
    skips.sort(key=lambda p: p.name)
    return memos + defaults + skips


def _scan_text(flat: str, patterns: list[re.Pattern], pdf_name: str, kind: str) -> list[dict]:
    """Run each pattern; return all matches with value + 200-char context."""
    out = []
    for pat in patterns:
        for m in pat.finditer(flat):
            val = _parse_pct(m.group(1))
            if val is None:
                continue
            ctx_start = max(0, m.start() - 80)
            ctx_end = min(len(flat), m.end() + 80)
            snippet = flat[ctx_start:ctx_end].strip()
            snippet = re.sub(r"\s+", " ", snippet)
            out.append({
                "kind": kind,
                "value": val,
                "snippet": snippet,
                "pdf": pdf_name,
            })
    return out


def _scan_pdf(pdf: Path) -> list[dict]:
    text, status = extract_pdf_text_with_timeout(pdf, timeout_s=PER_PDF_TIMEOUT_S)
    if status != "ok" or not text:
        return []
    flat = _normalize_text(text)
    matches = []
    matches += _scan_text(flat, OVERALL_IMPACT_PATTERNS, pdf.name, "overall_impact")
    matches += _scan_text(flat, BASE_RATE_PATTERNS, pdf.name, "base_rate")
    matches += _scan_text(flat, INDICATED_PATTERNS, pdf.name, "indicated")
    matches += _scan_text(flat, REQUESTED_PATTERNS, pdf.name, "requested")
    matches += _scan_text(flat, APPROVED_PATTERNS, pdf.name, "approved")
    return matches


def _scan_filing(state: str, filing_id: str) -> list[dict]:
    pdf_dir = PDF_ROOT / state / filing_id
    if not pdf_dir.exists():
        return []
    out = []
    for pdf in _prioritize(sorted(pdf_dir.glob("*.pdf"))):
        try:
            size_mb = pdf.stat().st_size / (1024 * 1024)
        except OSError:
            continue
        if size_mb > LARGE_PDF_SKIP_MB:
            continue
        out.extend(_scan_pdf(pdf))
    return out


def _classify(matches: list[dict], current_value: Optional[float]) -> dict:
    """Decide rate_change_type for one filing.

    Logic:
      - If overall_impact match exists AND its value is close to or differs
        from current → use 'overall_impact'. Suggest the overall value if
        different from current.
      - If only base_rate matches exist → 'base_rate'
      - If only indicated matches exist → 'indicated'
      - If multiple types present → 'overall_impact' if any overall, else
        prefer requested/approved over indicated/base.
      - If no matches at all → 'ambiguous' (PDF didn't expose any labeled type)
    """
    if not matches:
        return {"type": "ambiguous", "reason": "no_labeled_rate_phrase_found", "suggested_value": None}

    by_kind: dict[str, list[dict]] = defaultdict(list)
    for m in matches:
        by_kind[m["kind"]].append(m)

    # Pick a representative value per kind (first match — they're often duplicates)
    overall_vals = [m["value"] for m in by_kind.get("overall_impact", [])]
    base_vals = [m["value"] for m in by_kind.get("base_rate", [])]
    ind_vals = [m["value"] for m in by_kind.get("indicated", [])]
    req_vals = [m["value"] for m in by_kind.get("requested", [])]
    app_vals = [m["value"] for m in by_kind.get("approved", [])]

    # Approved wins if present
    if app_vals:
        return {
            "type": "overall_impact" if overall_vals else "approved",
            "reason": "approved_match",
            "suggested_value": app_vals[0],
        }
    if overall_vals:
        # Pick the overall value closest to current, if current is set
        if current_value is not None:
            best = min(overall_vals, key=lambda v: abs(v - current_value))
        else:
            best = overall_vals[0]
        return {"type": "overall_impact", "reason": "overall_impact_match", "suggested_value": best}
    if base_vals and not ind_vals and not req_vals:
        return {"type": "base_rate", "reason": "only_base_rate_found", "suggested_value": base_vals[0]}
    if ind_vals and not base_vals and not req_vals:
        return {"type": "indicated", "reason": "only_indicated_found", "suggested_value": ind_vals[0]}
    if req_vals and not base_vals and not ind_vals:
        # Requested without "overall" qualifier — could be either base or overall
        return {"type": "ambiguous", "reason": "requested_without_overall_qualifier", "suggested_value": req_vals[0]}
    # Mixed: prefer requested > base > indicated
    if req_vals:
        return {"type": "ambiguous", "reason": "mixed_types_no_overall", "suggested_value": req_vals[0]}
    if base_vals:
        return {"type": "base_rate", "reason": "base_rate_with_indicated", "suggested_value": base_vals[0]}
    return {"type": "indicated", "reason": "indicated_only", "suggested_value": ind_vals[0]}


def _load_targets() -> list[dict]:
    """Load the 23 rows from rate_changes.xlsx + look up filing_id from all_states."""
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Rate Changes"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    rate_rows = [dict(zip(header, r)) for r in rows[1:]]

    wb2 = openpyxl.load_workbook(ALL_STATES, read_only=True, data_only=True)
    ws2 = wb2["Filings"]
    rows2 = list(ws2.iter_rows(values_only=True))
    header2 = list(rows2[0])
    by_serff = {}
    for r in rows2[1:]:
        d = dict(zip(header2, r))
        serff = d.get("serff_tracking_number")
        if serff:
            by_serff[serff] = d

    out = []
    for r in rate_rows:
        serff = r.get("serff_tracking_number")
        full = by_serff.get(serff, {})
        out.append({
            "state": r.get("state"),
            "carrier": r.get("carrier"),
            "company_name": r.get("company_name"),
            "line_of_business": r.get("line_of_business"),
            "serff_tracking_number": serff,
            "filing_id": str(full.get("filing_id") or ""),
            "current_rate_effect_value": r.get("rate_effect_value"),
            "current_rate_effect_source": r.get("rate_effect_source"),
        })
    return out


def _parse_current_value(s) -> Optional[float]:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).replace("%", "").replace("+", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def main() -> int:
    targets = _load_targets()
    print(f"[load] {len(targets)} filings to audit\n")

    audit: list[dict] = []
    for i, t in enumerate(targets, 1):
        cur = _parse_current_value(t["current_rate_effect_value"])
        matches = _scan_filing(t["state"], t["filing_id"])
        cls = _classify(matches, cur)
        kinds_seen = sorted({m["kind"] for m in matches})
        print(
            f"  [{i:2d}/{len(targets)}] {t['serff_tracking_number']:22s} "
            f"current={t['current_rate_effect_value']:>9s}  "
            f"type={cls['type']:14s}  kinds_in_pdf={kinds_seen or '(none)'}"
        )
        audit.append({
            **t,
            "current_value_parsed": cur,
            "rate_change_type": cls["type"],
            "classification_reason": cls["reason"],
            "suggested_overall_value": cls["suggested_value"],
            "kinds_in_pdf": kinds_seen,
            "match_count": len(matches),
            "matches": matches,
        })

    AUDIT_JSON.write_text(json.dumps(audit, indent=2, default=str), encoding="utf-8")

    # ------ summary ------
    print("\n" + "=" * 60)
    print("AUDIT SUMMARY")
    print("=" * 60)
    type_counts = defaultdict(int)
    correctable = 0
    no_pdf_evidence = 0
    overall_matches_current = 0
    overall_differs_from_current = 0
    for a in audit:
        type_counts[a["rate_change_type"]] += 1
        if not a["matches"]:
            no_pdf_evidence += 1
        if a["rate_change_type"] == "overall_impact" and a["suggested_overall_value"] is not None:
            cur = a["current_value_parsed"]
            sug = a["suggested_overall_value"]
            if cur is not None and abs(cur - sug) < 0.01:
                overall_matches_current += 1
            else:
                overall_differs_from_current += 1
                # Correctable if current was a non-overall value and PDF reveals overall
                if "base_rate" in a["kinds_in_pdf"] or "indicated" in a["kinds_in_pdf"]:
                    correctable += 1

    print(f"\nClassification breakdown:")
    for kind in ("overall_impact", "base_rate", "indicated", "approved", "ambiguous"):
        print(f"  {kind:18s}  {type_counts.get(kind, 0)}")
    print(f"\nOf overall_impact rows:")
    print(f"  PDF overall == current value:    {overall_matches_current}")
    print(f"  PDF overall != current value:    {overall_differs_from_current}")
    print(f"  Of differing: also has base/indicated in PDF (correctable): {correctable}")
    print(f"\nFilings with NO labeled rate phrase in any PDF: {no_pdf_evidence}/{len(audit)}")

    # ------ ambiguous samples ------
    amb = [a for a in audit if a["rate_change_type"] == "ambiguous"]
    print(f"\nAMBIGUOUS CASES ({len(amb)}) — first 3 with snippets:\n")
    for a in amb[:3]:
        print(f"  {a['serff_tracking_number']} ({a['state']}, {a['carrier']}): "
              f"current={a['current_rate_effect_value']}  reason={a['classification_reason']}")
        if a["matches"]:
            for m in a["matches"][:3]:
                print(f"    [{m['kind']}] {m['value']:+.2f}%  ({m['pdf']})")
                print(f"       \"...{m['snippet'][:160]}...\"")
        else:
            print(f"    (no labeled rate phrase found in any PDF — only the SERFF table value is available)")
        print()

    print(f"[write] {AUDIT_JSON}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
