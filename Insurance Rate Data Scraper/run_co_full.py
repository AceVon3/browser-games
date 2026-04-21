"""Full CO enrichment run (Step 8): all target-company filings in Colorado.

Mirrors `run_id_full.py` with CO-specific paths. Loads the search-only workbook
(`output/co_all_companies_search.xlsx`), hydrates Filing objects, then runs
detail-page enrichment group-by-group. Rate/Rule filings get PDF download +
parse; Form and Scoring Model filings get metadata only.

Final output: `output/co_final.xlsx` plus a printed summary.
"""
from __future__ import annotations

import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from src.config import HEADLESS, OUTPUT_DIR, PDF_DIR, REQUEST_DELAY, USER_AGENT
from src.detail import enrich_filing
from src.models import Filing
from src.output import _is_unparseable, write_excel
from src.search import _parse_date, _set_rows_per_page_100, _submit_search


def _parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return _parse_date(s)


SEARCH_XLSX = OUTPUT_DIR / "co_all_companies_search.xlsx"
FINAL_XLSX = OUTPUT_DIR / "co_final.xlsx"
DEFAULT_STATE = "CO"


def _load_filings(path: Path) -> list[Filing]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {name: i for i, name in enumerate(header)}

    def g(r, col):
        i = idx.get(col)
        return r[i] if i is not None else None

    filings: list[Filing] = []
    for r in rows[1:]:
        naic_raw = g(r, "naic_codes") or ""
        naic = [c.strip() for c in str(naic_raw).split(";") if c.strip()]
        sub_date = _parse_any_date(g(r, "submission_date"))
        filings.append(
            Filing(
                state=g(r, "state") or DEFAULT_STATE,
                serff_tracking_number=g(r, "serff_tracking_number") or "",
                filing_id=str(g(r, "filing_id") or ""),
                company_name=g(r, "company_name") or "",
                target_company=g(r, "target_company") or "",
                naic_codes=naic,
                product_name=g(r, "product_name"),
                type_of_insurance=g(r, "type_of_insurance"),
                sub_type_of_insurance=g(r, "sub_type_of_insurance"),
                filing_type=g(r, "filing_type"),
                filing_status=g(r, "filing_status"),
                submission_date=sub_date,
                detail_url=g(r, "detail_url"),
            )
        )
    return filings


def _summarize(filings: list[Filing], elapsed_s: float) -> None:
    rate_rule = [f for f in filings if f.filing_type == "Rate/Rule"]
    nonzero = [
        f for f in rate_rule
        if (f.overall_rate_effect not in (None, 0.0))
        or (f.requested_rate_effect not in (None, 0.0))
        or (f.approved_rate_effect not in (None, 0.0))
    ]
    premium_neutral = [f for f in rate_rule if f.overall_rate_effect == 0.0]
    unparseable = [f for f in rate_rule if _is_unparseable(f)]

    pdf_bytes = 0
    pdf_count = 0
    download_fails = 0
    for f in rate_rule:
        pdf_count += len(f.pdfs)
        for p in f.pdfs:
            if p.local_path:
                pth = Path(p.local_path)
                if pth.exists():
                    pdf_bytes += pth.stat().st_size
                else:
                    download_fails += 1
        if f.pdf_parse_status == "no_pdfs_downloaded":
            download_fails += 1

    metadata_only = [f for f in filings if f.filing_type != "Rate/Rule"]
    backfilled = [f for f in filings if f.submission_date is not None]
    skipped_no_date = [f for f in filings if f.submission_date is None]

    print("\n" + "=" * 60)
    print(f"CO FULL RUN SUMMARY ({len(filings)} filings)")
    print("=" * 60)
    print(f"Runtime:                      {elapsed_s/60:.1f} min ({elapsed_s:.0f}s)")
    print(f"Rate/Rule filings:            {len(rate_rule)}")
    print(f"  extracted rate effects:     {len(nonzero)}")
    print(f"  premium-neutral (0.0%):     {len(premium_neutral)}")
    print(f"  unparseable (manual queue): {len(unparseable)}")
    print(f"Non-Rate/Rule (metadata only): {len(metadata_only)}")
    print(f"Submission dates populated:   {len(backfilled)}/{len(filings)}")
    if skipped_no_date:
        print(f"  still missing dates:        {len(skipped_no_date)}")
    print(f"PDFs downloaded:              {pdf_count}")
    print(f"PDF storage used:             {pdf_bytes/(1024*1024):.1f} MB")
    print(f"PDF download failures:        {download_fails}")

    if nonzero:
        print("\nExtracted rate effects:")
        for f in nonzero:
            pct = f.overall_rate_effect if f.overall_rate_effect is not None else (
                f.approved_rate_effect if f.approved_rate_effect is not None else f.requested_rate_effect
            )
            print(f"  {f.serff_tracking_number:22s} {f.target_company:15s} {pct:+.2f}%")

    if unparseable:
        print(f"\nUnparseable queue ({len(unparseable)} filings — manual review needed):")
        for f in unparseable:
            print(f"  {f.serff_tracking_number:22s} {f.target_company:15s} pdfs={len(f.pdfs)} status={f.pdf_parse_status}")

    no_pdfs = [
        f for f in rate_rule
        if not f.pdfs or f.pdf_parse_status in ("no_pdfs_attached", "no_pdfs_downloaded")
    ]
    print(f"\nRate/Rule filings with no PDFs attached/downloaded: {len(no_pdfs)}")

    print("\nParse rates by carrier (Rate/Rule only):")
    print(f"  {'carrier':15s}  {'total':>5s}  {'parsed':>6s}  {'neutral':>7s}  {'unparse':>7s}  {'no_pdf':>6s}  rate")
    by_carrier: dict[str, list[Filing]] = defaultdict(list)
    for f in rate_rule:
        by_carrier[f.target_company].append(f)
    for carrier in sorted(by_carrier):
        group = by_carrier[carrier]
        n = len(group)
        n_parsed = sum(
            1 for f in group
            if (f.overall_rate_effect not in (None,)) or
               (f.requested_rate_effect is not None) or
               (f.approved_rate_effect is not None)
        )
        n_neutral = sum(1 for f in group if f.overall_rate_effect == 0.0)
        n_unparse = sum(1 for f in group if _is_unparseable(f))
        n_nopdf = sum(
            1 for f in group
            if not f.pdfs or f.pdf_parse_status in ("no_pdfs_attached", "no_pdfs_downloaded")
        )
        pct = f"{100.0 * n_parsed / n:.1f}%" if n else "-"
        print(f"  {carrier:15s}  {n:>5d}  {n_parsed:>6d}  {n_neutral:>7d}  {n_unparse:>7d}  {n_nopdf:>6d}  {pct}")


def main() -> int:
    print(f"[load] {SEARCH_XLSX}", flush=True)
    filings = _load_filings(SEARCH_XLSX)
    print(f"[load] {len(filings)} filings hydrated", flush=True)

    groups: dict[tuple[str, str], list[Filing]] = defaultdict(list)
    for f in filings:
        groups[(f.state, f.target_company)].append(f)

    started = time.time()
    group_order = sorted(groups.keys(), key=lambda k: len(groups[k]))

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        try:
            for gi, (state, company) in enumerate(group_order, 1):
                group = groups[(state, company)]
                print(
                    f"\n[group {gi}/{len(group_order)}] {state} / {company}: "
                    f"{len(group)} filings ({sum(1 for f in group if f.filing_type=='Rate/Rule')} Rate/Rule)",
                    flush=True,
                )
                ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
                page = ctx.new_page()
                try:
                    if not _submit_search(page, state, company):
                        print(f"  ! search failed, skipping {company}", flush=True)
                        continue
                    _set_rows_per_page_100(page)

                    for i, f in enumerate(group, 1):
                        enrich_filing(page, f, download_pdfs=True)
                        if i % 10 == 0 or i == len(group):
                            elapsed = (time.time() - started) / 60
                            print(
                                f"    [checkpoint] {i}/{len(group)} enriched "
                                f"(total elapsed: {elapsed:.1f}m)",
                                flush=True,
                            )
                        time.sleep(REQUEST_DELAY)
                finally:
                    ctx.close()

                write_excel(filings, FINAL_XLSX)
                print(f"  [save] {FINAL_XLSX}", flush=True)
        finally:
            browser.close()

    elapsed_s = time.time() - started
    write_excel(filings, FINAL_XLSX)
    _summarize(filings, elapsed_s)
    print(f"\nFinal output: {FINAL_XLSX}")
    print(f"PDF archive:  {PDF_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
