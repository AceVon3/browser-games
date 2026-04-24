"""Build the canonical multi-state final-rates deliverable.

Concatenates ID + WA + CO + OR per-state workbooks into:
  - output/all_states_final_rates.xlsx  (rate_filings sheet + README sheet)
  - output/all_states_final_rates.csv   (rate_filings only, for quick load)
"""
from pathlib import Path
import csv
import openpyxl
from openpyxl.styles import Font, Alignment

STATES = ["ID", "WA", "CO", "OR"]
OUT_XLSX = Path("output/all_states_final_rates.xlsx")
OUT_CSV = Path("output/all_states_final_rates.csv")

README_LINES: list[tuple[str, str]] = [
    ("Insurance Rate Filings — Three-State Dataset", ""),
    ("", ""),
    ("Source", "SERFF Filing Access (filingaccess.serff.com), system-generated Filing Summary PDF"),
    ("Pipeline", "Public search -> minimal zip download -> {tracking}.pdf -> Disposition / Company Rate Information table"),
    ("States", "Idaho (ID), Washington (WA), Colorado (CO), Oregon (OR)"),
    ("Lines", "19.0 Personal Auto, 04.0 Homeowners, 03.0 Personal Farmowners (NAIC TOI codes)"),
    ("Carriers", "State Farm, GEICO, Progressive, Allstate, Travelers, Liberty Mutual (and named subsidiaries); plus Safeco (Liberty Mutual independent-agent brand) and Encompass (Allstate independent-agent brand) — searched separately because each files under its own brand on SERFF"),
    ("Filing types kept", "Rate, Rate/Rule (Form-only and Rule-only filings excluded)"),
    ("Excluded carriers", "Drive Insurance (Progressive, retired); Esurance (Allstate, wound down 2020); United Financial and other niche specialty subsidiaries"),
    ("Excluded filings", "New-program / new-product / 'Introduction of' filings; filings flagged 'Rate data does NOT apply to filing.' by the filer"),
    ("", ""),
    ("Anchor validation", "Idaho SFMA-134676753 matches AM Best Disposition Page Data on all 14 fields"),
    ("Format", "Columns ordered to match AM Best Disposition Page Data export"),
    ("", ""),
    ("FIELD DEFINITIONS", ""),
    ("state", "Two-letter state code"),
    ("effective_date", "Requested effective date (Renewal preferred over New)"),
    ("company_name", "Subsidiary writing the rate; for multi-company filings expanded one row per subsidiary"),
    ("line_of_business", "NAIC parent TOI code + label, e.g. '19.0 Personal Auto' (kept for AM Best compatibility)"),
    ("sub_type_of_insurance", "NAIC Sub-TOI code + label, e.g. '19.0001 Private Passenger Auto (PPA)' / '19.0002 Motorcycle' / '19.0003 RV'"),
    ("overall_indicated_change", "Filer's actuarially indicated rate change (may be blank when filer omits)"),
    ("overall_rate_impact", "Filed rate impact (the change actually requested)"),
    ("written_premium_change", "Effect of rate filing on written premium, USD"),
    ("policyholders_affected", "Count of policyholders impacted"),
    ("written_premium_for_program", "Total written premium for the program, USD"),
    ("maximum_percent_change", "Largest individual policyholder increase under the filing"),
    ("minimum_percent_change", "Largest individual policyholder decrease (most negative) under the filing"),
    ("rate_activity", "rate_change | rate_change_withdrawn | rate_change_disapproved | rate_change_pending"),
    ("serff_tracking_number", "SERFF filing tracking number (carrier-prefixed)"),
    ("disposition_status", "State decision: Approved/Filed/Withdrawn/Disapproved/Pending (case as filed)"),
    ("filing_date", "Date the filing was submitted to the state"),
    ("source_pdf", "Relative path to the cached system-generated Filing Summary PDF"),
    ("", ""),
    ("LIMITATIONS", ""),
    ("Date range", "Filings present in SERFF Public Access at run time (no explicit date filter)"),
    ("Carrier scope", "Only the six listed national groups + their named subsidiaries"),
    ("Line scope", "Personal Auto, Homeowners, Farmowners only"),
    ("State scope", "ID, WA, CO, OR only"),
    ("Filer flag", "Rows excluded when filer set 'Rate data does NOT apply to filing.'; this flag is taken at face value"),
    ("PDF parsing", "Three Disposition row patterns supported (full / blank-indicated / sparse). Layouts outside these patterns may be missed"),
    ("", ""),
    ("RECOMMENDED USE", ""),
    ("", "Comparative analysis of approved/filed rate changes across ID/WA/CO/OR for the named carriers"),
    ("", "Cross-reference to AM Best Disposition Page Data using serff_tracking_number"),
    ("", "Not a substitute for full-state market analysis — scope is bounded by the carrier and line filters above"),
]


def write_workbook(rows_by_state: dict[str, list[list]], header: list[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "rate_filings"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for state in STATES:
        for r in rows_by_state[state]:
            ws.append(r)
    ws.freeze_panes = "A2"
    for col_letter in [chr(c) for c in range(ord("A"), ord("A") + len(header))]:
        ws.column_dimensions[col_letter].width = 18

    rd = wb.create_sheet("README")
    rd.column_dimensions["A"].width = 32
    rd.column_dimensions["B"].width = 110
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for label, value in README_LINES:
        rd.append([label, value])
        row = rd[rd.max_row]
        row[0].font = bold
        row[1].alignment = wrap

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_XLSX))


def write_csv(rows_by_state: dict[str, list[list]], header: list[str]) -> None:
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for state in STATES:
            for r in rows_by_state[state]:
                w.writerow(["" if v is None else v for v in r])


def main() -> None:
    src_paths = [Path(f"output/{s.lower()}_final_rates.xlsx") for s in STATES]
    for p in src_paths:
        if not p.exists():
            raise SystemExit(f"missing input: {p}")

    header: list[str] | None = None
    rows_by_state: dict[str, list[list]] = {}
    for state, path in zip(STATES, src_paths):
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        hdr = list(next(rows))
        if header is None:
            header = hdr
        elif hdr != header:
            raise SystemExit(f"header mismatch in {path}: {hdr}")
        rows_by_state[state] = [list(r) for r in rows]
        wb.close()

    write_workbook(rows_by_state, header)
    write_csv(rows_by_state, header)

    total = sum(len(v) for v in rows_by_state.values())
    print(f"wrote {OUT_XLSX} (rate_filings + README sheet)")
    print(f"wrote {OUT_CSV}")
    print(f"  total: {total}")
    for s in STATES:
        print(f"  {s}: {len(rows_by_state[s])}")


if __name__ == "__main__":
    main()
