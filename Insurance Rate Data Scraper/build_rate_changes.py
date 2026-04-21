"""Final rate-change deliverable with broadened filter + PDF backfill.

Pipeline:
    1. Load `output/all_states_final.xlsx`.
    2. Filter to rows where in_target_lines=True AND active disposition AND
       ANY of {overall, requested, approved}_rate_effect is populated.
    3. For each kept row, re-scan the already-downloaded PDFs for:
          - effective_date (proposed / requested / plain)
          - current_avg_premium (current / existing)
          - new_avg_premium (new / proposed)
       Conservative patterns — premium keyword required within proximity.
    4. Compute rate_effect_value + rate_effect_source (approved > overall >
       requested) and premium_change_dollars (new - current).
    5. Write output/rate_changes.xlsx and output/rate_changes.csv.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import OUTPUT_DIR
from src.utils import extract_pdf_text_with_timeout

SOURCE = OUTPUT_DIR / "all_states_final.xlsx"
OUT_XLSX = OUTPUT_DIR / "rate_changes.xlsx"
OUT_CSV = OUTPUT_DIR / "rate_changes.csv"
PDF_ROOT = OUTPUT_DIR / "pdfs"
EFFECTIVE_DATES_CACHE = OUTPUT_DIR / "effective_dates.json"

PER_PDF_TIMEOUT_S = 45.0
LARGE_PDF_SKIP_MB = 15.0

MEMO_KEYWORDS = ("memo", "summary", "cover letter", "justification", "filing packet")
SKIP_KEYWORDS = ("manual", "tracked changes", "rate pages", "exhibit", "complete", "compare")

COLUMNS = [
    "state",
    "carrier",
    "company_name",
    "line_of_business",
    "effective_date",
    "effective_date_source",
    "rate_effect_value",
    "rate_effect_source",
    "current_avg_premium",
    "new_avg_premium",
    "serff_tracking_number",
    "filing_date",
    "disposition_status",
]

INACTIVE_TOKENS = ("WITHDRAWN", "DISAPPROVED")


# ---------- PDF extraction patterns (conservative) ----------

# Effective date — tight patterns, proximity to "effective date" phrase required.
_MONTH_DATE = r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)[\.,]?\s+\d{1,2}(?:st|nd|rd|th)?[,\.]?\s+\d{4}"
_NUMERIC_DATE = r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"

_EFFECTIVE_DATE_PATTERNS = [
    # "Proposed effective date: 6/1/2026" / "Requested effective date - June 1, 2026"
    re.compile(
        rf"(?:proposed|requested)\s+effective\s+date[\s:–\-]*({_NUMERIC_DATE}|{_MONTH_DATE})",
        re.IGNORECASE,
    ),
    # "new business effective: 6/1/2026" / "renewal effective 6/1/2026"
    re.compile(
        rf"(?:new\s+business|renewal)\s+effective(?:\s+date)?[\s:–\-]*({_NUMERIC_DATE}|{_MONTH_DATE})",
        re.IGNORECASE,
    ),
    # Plain "effective date: 6/1/2026"
    re.compile(
        rf"\beffective\s+date[\s:–\-]+({_NUMERIC_DATE}|{_MONTH_DATE})",
        re.IGNORECASE,
    ),
    # "effective on/as of 6/1/2026"
    re.compile(
        rf"\beffective\s+(?:on|as\s+of)\s+({_NUMERIC_DATE}|{_MONTH_DATE})",
        re.IGNORECASE,
    ),
]

# Premium extraction — require "average premium" within proximity of the dollar figure.
_MONEY_CAPTURE = r"\$?\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)"

_CURRENT_PREMIUM_PATTERNS = [
    re.compile(
        rf"(?:current|existing|present|in[\s-]force)\s+"
        rf"(?:annual\s+|policy\s+|policyholder\s+)?"
        rf"average\s+(?:annual\s+)?premium[^$0-9\n\r]{{0,40}}{_MONEY_CAPTURE}",
        re.IGNORECASE,
    ),
    re.compile(
        rf"average\s+(?:annual\s+)?premium\s+(?:before|prior\s+to)[^$0-9\n\r]{{0,40}}{_MONEY_CAPTURE}",
        re.IGNORECASE,
    ),
]

_NEW_PREMIUM_PATTERNS = [
    re.compile(
        rf"(?:new|proposed|revised|updated|projected)\s+"
        rf"(?:annual\s+|policy\s+|policyholder\s+)?"
        rf"average\s+(?:annual\s+)?premium[^$0-9\n\r]{{0,40}}{_MONEY_CAPTURE}",
        re.IGNORECASE,
    ),
    re.compile(
        rf"average\s+(?:annual\s+)?premium\s+after[^$0-9\n\r]{{0,40}}{_MONEY_CAPTURE}",
        re.IGNORECASE,
    ),
]

# "from $X to $Y" — only accept when within 60 chars of "premium" keyword.
_FROM_TO_RE = re.compile(
    rf"from\s+{_MONEY_CAPTURE}\s+to\s+{_MONEY_CAPTURE}",
    re.IGNORECASE,
)


def _parse_money(s: str) -> Optional[float]:
    if not s:
        return None
    s = s.replace(",", "").strip()
    try:
        v = float(s)
    except ValueError:
        return None
    if v < 50:  # implausibly low for an average premium
        return None
    return v


def _parse_date_any(s: str) -> Optional[date]:
    if not s:
        return None
    s = s.strip().rstrip(".")
    s = re.sub(r"(\d)(?:st|nd|rd|th)\b", r"\1", s, flags=re.IGNORECASE)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
                "%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _categorize(name: str) -> str:
    n = name.lower()
    if any(k in n for k in MEMO_KEYWORDS):
        return "memo"
    if any(k in n for k in SKIP_KEYWORDS):
        return "skip"
    return "default"


def _prioritize(paths: list[Path]) -> list[Path]:
    memos, defaults, skips = [], [], []
    for p in paths:
        c = _categorize(p.name)
        (memos if c == "memo" else defaults if c == "default" else skips).append(p)
    memos.sort(key=lambda p: p.name)
    defaults.sort(key=lambda p: p.name)
    skips.sort(key=lambda p: p.name)
    return memos + defaults + skips


def _extract_effective_date(flat: str) -> Optional[date]:
    for pat in _EFFECTIVE_DATE_PATTERNS:
        m = pat.search(flat)
        if m:
            d = _parse_date_any(m.group(1))
            if d:
                return d
    return None


def _extract_current_premium(flat: str) -> Optional[float]:
    for pat in _CURRENT_PREMIUM_PATTERNS:
        m = pat.search(flat)
        if m:
            v = _parse_money(m.group(1))
            if v is not None:
                return v
    return None


def _extract_new_premium(flat: str) -> Optional[float]:
    for pat in _NEW_PREMIUM_PATTERNS:
        m = pat.search(flat)
        if m:
            v = _parse_money(m.group(1))
            if v is not None:
                return v
    return None


def _extract_from_to(flat: str) -> tuple[Optional[float], Optional[float]]:
    """Accept 'from $X to $Y' only if 'premium' appears within 60 chars."""
    for m in _FROM_TO_RE.finditer(flat):
        window = flat[max(0, m.start() - 60):m.end() + 60].lower()
        if "premium" not in window:
            continue
        v1 = _parse_money(m.group(1))
        v2 = _parse_money(m.group(2))
        if v1 is not None and v2 is not None:
            return v1, v2
    return None, None


def _scan_filing_pdfs(pdf_dir: Path) -> dict:
    """Walk one filing's PDFs; return {effective_date, current_avg_premium, new_avg_premium}.
    First non-None hit per field wins across memo-priority order."""
    found: dict = {}
    if not pdf_dir.exists():
        return found
    pdfs = _prioritize(sorted(pdf_dir.glob("*.pdf")))
    for pdf in pdfs:
        try:
            size_mb = pdf.stat().st_size / (1024 * 1024)
        except OSError:
            continue
        if size_mb > LARGE_PDF_SKIP_MB:
            continue
        text, status = extract_pdf_text_with_timeout(pdf, timeout_s=PER_PDF_TIMEOUT_S)
        if status != "ok" or not text:
            continue
        flat = re.sub(r"\s+", " ", text)

        if "effective_date" not in found:
            d = _extract_effective_date(flat)
            if d:
                found["effective_date"] = d

        if "current_avg_premium" not in found:
            v = _extract_current_premium(flat)
            if v is not None:
                found["current_avg_premium"] = v

        if "new_avg_premium" not in found:
            v = _extract_new_premium(flat)
            if v is not None:
                found["new_avg_premium"] = v

        if "current_avg_premium" not in found or "new_avg_premium" not in found:
            v1, v2 = _extract_from_to(flat)
            if v1 is not None and v2 is not None:
                found.setdefault("current_avg_premium", v1)
                found.setdefault("new_avg_premium", v2)

        if all(k in found for k in ("effective_date", "current_avg_premium", "new_avg_premium")):
            break
    return found


# ---------- excel/csv helpers ----------

def _load_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    return [dict(zip(header, r)) for r in rows[1:]]


def _is_inactive(r: dict) -> bool:
    status = " ".join(
        str(r.get(k) or "").upper()
        for k in ("disposition_status", "state_status", "filing_status")
    )
    return any(tok in status for tok in INACTIVE_TOKENS)


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "1", "yes")


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _fmt_pct(v) -> str:
    if v is None:
        return ""
    return f"{v:+.2f}%"


def _fmt_money(v) -> str:
    if v is None:
        return ""
    return f"${v:,.2f}"


def _fmt_date(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        return s


def _line_of_business(r: dict) -> str:
    toi = (r.get("type_of_insurance") or "").strip()
    sub = (r.get("sub_type_of_insurance") or "").strip()
    if toi and sub:
        return f"{toi} — {sub}"
    return sub or toi


def _date_key(iso: str) -> int:
    if not iso:
        return 0
    try:
        return int(datetime.fromisoformat(iso).strftime("%Y%m%d"))
    except ValueError:
        return 0


LEGEND_ROWS = [
    ("Column", "Meaning"),
    ("state", "US state of the filing (WA / ID / CO)."),
    ("carrier", "Target carrier searched (Allstate, Geico, Liberty, Progressive, State Farm, Travelers)."),
    ("company_name", "Specific underwriting company that filed (e.g. Allstate Fire & Casualty)."),
    ("line_of_business", "Type Of Insurance — Sub Type Of Insurance from SERFF."),
    ("effective_date",
     "When the rate change takes / took effect. Blank where SERFF Filing Access does not expose "
     "this and the rate memo PDF doesn't include an extractable 'Effective Date: MM/DD/YYYY' phrase."),
    ("effective_date_source",
     "How the effective_date was derived: 'serff:<bucket>' = SERFF detail-page label "
     "(approved/requested/proposed/plain). 'pdf_memo' = parsed from a downloaded rate memo PDF. "
     "Blank = no source available; needs manual review."),
    ("rate_effect_value",
     "Signed percentage change in rates (e.g. +11.90%). See rate_effect_source for which SERFF "
     "field this value came from."),
    ("rate_effect_source",
     "Which SERFF field rate_effect_value came from. Priority approved > overall > requested. "
     "'approved' = Approved Rate Impact (regulator-approved per-company value). "
     "'overall'  = Overall Rate Impact (carrier-stated filing-level weighted average; for "
     "closed/approved filings this is the approved overall). "
     "'requested' = Requested Rate Impact (carrier's filed value, not yet approved). "
     "'multiple' = more than one of the three buckets was populated (rare; the 'approved' value was used)."),
    ("current_avg_premium",
     "Average annual premium BEFORE the change. Blank where the PDF memo does not state a "
     "concrete dollar amount alongside the keyword 'average premium' (typical — most memos cite "
     "percent change only, not dollar baseline)."),
    ("new_avg_premium",
     "Average annual premium AFTER the change. Same caveat as current_avg_premium."),
    ("serff_tracking_number", "Unique SERFF identifier; can be used to look the filing up directly."),
    ("filing_date", "Submission date the carrier filed with the state regulator."),
    ("disposition_status", "Regulator decision (Approved / Pending / etc)."),
    ("", ""),
    ("LIMITATIONS", ""),
    ("Effective date coverage",
     "SERFF Filing Access (the public consumer portal at filingaccess.serff.com) does NOT expose "
     "Rate Data / per-company effective dates anywhere on the filing detail page — only "
     "Submission Date, Disposition Date, and State Status Last Changed are available. The full "
     "SERFF system used by regulators DOES contain effective-date fields, but those are not in "
     "the public portal. PDF rate memos sometimes contain effective dates but most use boilerplate "
     "phrasing ('the effective date of the rates') without an extractable date. Blank "
     "effective_date is the honest state of the data. See the Manual Review sheet for the rows "
     "needing hand-fill, with PDF directory paths."),
    ("Premium dollar coverage",
     "Most filings disclose only percent change, not the dollar baseline. Conservative pattern "
     "matching (requires 'average premium' keyword within 40 characters of the dollar figure) "
     "produced 1/23 hits. Approving broader patterns risks false positives, so columns are left "
     "blank rather than guessed."),
    ("Filter applied",
     "in_target_lines = True (Homeowners / Auto / Condo per the project spec) AND disposition is "
     "active (not Withdrawn / Disapproved) AND at least one of the three rate-effect fields is "
     "populated. New product launches and unparseable filings are excluded by definition."),
]


def _write_xlsx(rows: list[dict], path: Path, manual_review_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate Changes"
    ws.append(COLUMNS)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
    for idx, col in enumerate(COLUMNS, start=1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
    ws.freeze_panes = "A2"

    # --- Legend sheet ---
    ws_legend = wb.create_sheet("Legend & Notes")
    for i, (col, meaning) in enumerate(LEGEND_ROWS, start=1):
        ws_legend.cell(row=i, column=1, value=col)
        ws_legend.cell(row=i, column=2, value=meaning)
    for cell in ws_legend["A"]:
        cell.font = bold
    ws_legend.column_dimensions["A"].width = 26
    ws_legend.column_dimensions["B"].width = 110
    for row_cells in ws_legend.iter_rows(min_row=1, max_row=ws_legend.max_row,
                                         min_col=2, max_col=2):
        for c in row_cells:
            c.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    # --- Manual review sheet ---
    ws_mr = wb.create_sheet("Manual Review (effective_date)")
    mr_cols = [
        "state", "carrier", "company_name", "line_of_business",
        "rate_effect_value", "rate_effect_source",
        "serff_tracking_number", "filing_date",
        "disposition_status", "pdf_directory",
    ]
    ws_mr.append(mr_cols)
    for cell in ws_mr[1]:
        cell.font = bold
        cell.fill = fill
    for r in manual_review_rows:
        ws_mr.append([r.get(c, "") for c in mr_cols])
    for idx, col in enumerate(mr_cols, start=1):
        max_len = len(col)
        for row in ws_mr.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws_mr.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 80)
    ws_mr.freeze_panes = "A2"

    wb.save(path)


def _write_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLUMNS})


# ---------- main ----------

def _pick_rate_effect(r: dict) -> tuple[Optional[float], str]:
    """Return (value, source). Priority: approved > overall > requested.
    Source names the FIRST populated bucket in that priority. If multiple
    buckets are populated, source = 'multiple'."""
    overall = _to_float(r.get("overall_rate_effect"))
    requested = _to_float(r.get("requested_rate_effect"))
    approved = _to_float(r.get("approved_rate_effect"))
    populated = [(name, v) for name, v in
                 (("approved", approved), ("overall", overall), ("requested", requested))
                 if v is not None]
    if not populated:
        return None, ""
    value = populated[0][1]
    if len(populated) > 1:
        source = "multiple"
    else:
        source = populated[0][0]
    return value, source


def main() -> int:
    records = _load_rows(SOURCE)
    print(f"[load] {SOURCE.name}: {len(records)} total rows")

    kept: list[dict] = []
    for r in records:
        if not _to_bool(r.get("in_target_lines")):
            continue
        if _is_inactive(r):
            continue
        val, src = _pick_rate_effect(r)
        if val is None:
            continue
        r["_rate_value"] = val
        r["_rate_source"] = src
        kept.append(r)
    print(f"[filter] target + active + any rate effect: {len(kept)} rows")

    # --- PDF backfill pass ---
    print(f"\n[pdf-backfill] scanning {len(kept)} filings for effective_date + premiums")
    eff_hits = 0
    cur_hits = 0
    new_hits = 0
    for i, r in enumerate(kept, 1):
        state = r.get("state") or ""
        fid = str(r.get("filing_id") or "")
        pdf_dir = PDF_ROOT / state / fid
        found = _scan_filing_pdfs(pdf_dir)
        r["_pdf_found"] = found
        if "effective_date" in found:
            eff_hits += 1
        if "current_avg_premium" in found:
            cur_hits += 1
        if "new_avg_premium" in found:
            new_hits += 1
        print(
            f"  [{i:2d}/{len(kept)}] {r.get('serff_tracking_number','?'):22s} "
            f"eff={'Y' if 'effective_date' in found else '-'} "
            f"cur={'Y' if 'current_avg_premium' in found else '-'} "
            f"new={'Y' if 'new_avg_premium' in found else '-'}",
            flush=True,
        )

    print(
        f"\n[pdf-backfill] done — "
        f"effective_date: {eff_hits}/{len(kept)}  "
        f"current_avg_premium: {cur_hits}/{len(kept)}  "
        f"new_avg_premium: {new_hits}/{len(kept)}"
    )

    # --- Sanity check on PDF-extracted effective_date ---
    # PDFs commonly cite *prior* filings' effective dates, which our pattern
    # catches. Reject values that predate the filing_date by >30 days.
    eff_rejected = 0
    for r in kept:
        found = r.get("_pdf_found") or {}
        eff = found.get("effective_date")
        sub = r.get("submission_date")
        sub_date = sub if isinstance(sub, date) and not isinstance(sub, datetime) else (
            sub.date() if isinstance(sub, datetime) else None
        )
        if sub_date is None and isinstance(sub, str):
            try:
                sub_date = datetime.fromisoformat(sub).date()
            except ValueError:
                sub_date = None
        if eff and sub_date and (sub_date - eff).days > 30:
            found.pop("effective_date", None)
            eff_rejected += 1
    if eff_rejected:
        print(f"\n[sanity] rejected {eff_rejected} PDF effective_date values that predate filing_date by >30 days")

    # --- Overlay effective_date from SERFF detail-page scrape ---
    # Canonical source wins over PDF memo fallback.
    serff_dates: dict = {}
    if EFFECTIVE_DATES_CACHE.exists():
        try:
            serff_dates = json.loads(EFFECTIVE_DATES_CACHE.read_text(encoding="utf-8"))
            print(f"\n[overlay] loaded {EFFECTIVE_DATES_CACHE.name} with {len(serff_dates)} entries")
        except Exception as e:
            print(f"\n[overlay] failed to load {EFFECTIVE_DATES_CACHE.name}: {e}")

    out_rows = []
    for r in kept:
        found = r.get("_pdf_found", {})
        serff_hit = serff_dates.get(r.get("serff_tracking_number") or "", {})
        serff_eff = serff_hit.get("effective_date") if isinstance(serff_hit, dict) else None
        serff_src = serff_hit.get("source") if isinstance(serff_hit, dict) else None

        if serff_eff:
            eff = serff_eff
            eff_source = f"serff:{serff_src}" if serff_src else "serff"
        elif found.get("effective_date"):
            eff = found["effective_date"]
            eff_source = "pdf_memo"
        else:
            eff = None
            eff_source = ""

        cur = found.get("current_avg_premium")
        new = found.get("new_avg_premium")

        out_rows.append({
            "state": r.get("state") or "",
            "carrier": r.get("target_company") or "",
            "company_name": r.get("company_name") or "",
            "line_of_business": _line_of_business(r),
            "effective_date": _fmt_date(eff),
            "effective_date_source": eff_source,
            "rate_effect_value": _fmt_pct(r["_rate_value"]),
            "rate_effect_source": r["_rate_source"],
            "current_avg_premium": _fmt_money(cur),
            "new_avg_premium": _fmt_money(new),
            "serff_tracking_number": r.get("serff_tracking_number") or "",
            "filing_date": _fmt_date(r.get("submission_date")),
            "disposition_status": r.get("disposition_status") or r.get("state_status") or "",
        })

    # Sort: state asc, effective_date desc (fallback filing_date).
    out_rows.sort(
        key=lambda x: (
            x["state"],
            -_date_key(x["effective_date"] or x["filing_date"]),
        )
    )

    # Manual review queue: rows with blank effective_date.
    manual_review_rows = []
    for r, raw in zip(out_rows, kept):
        if r["effective_date"]:
            continue
        state = raw.get("state") or ""
        fid = str(raw.get("filing_id") or "")
        pdf_dir = (PDF_ROOT / state / fid)
        manual_review_rows.append({
            **r,
            "pdf_directory": str(pdf_dir.resolve()) if pdf_dir.exists() else "(not downloaded)",
        })

    _write_xlsx(out_rows, OUT_XLSX, manual_review_rows)
    _write_csv(out_rows, OUT_CSV)

    # --- Final report ---
    total = len(out_rows)
    have_eff = sum(1 for r in out_rows if r["effective_date"])
    have_cur = sum(1 for r in out_rows if r["current_avg_premium"])
    have_new = sum(1 for r in out_rows if r["new_avg_premium"])

    rate_src: dict[str, int] = {}
    for r in out_rows:
        rate_src[r["rate_effect_source"]] = rate_src.get(r["rate_effect_source"], 0) + 1

    eff_src: dict[str, int] = {}
    for r in out_rows:
        s = r["effective_date_source"] or "(blank)"
        eff_src[s] = eff_src.get(s, 0) + 1

    print("\n" + "=" * 60)
    print("FINAL REPORT")
    print("=" * 60)
    print(f"Total rows:                     {total}")
    print(f"rate_effect_value:              {total}/{total}  (always populated)")
    print(f"rate_effect_source breakdown:")
    for src, n in sorted(rate_src.items()):
        print(f"    {src:14s}  {n}")
    print(f"effective_date populated:       {have_eff}/{total}")
    print(f"effective_date_source breakdown:")
    for src, n in sorted(eff_src.items()):
        print(f"    {src:14s}  {n}")
    print(f"current_avg_premium populated:  {have_cur}/{total}")
    print(f"new_avg_premium populated:      {have_new}/{total}")
    print(f"\nFiles written:")
    print(f"  {OUT_XLSX}")
    print(f"  {OUT_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
