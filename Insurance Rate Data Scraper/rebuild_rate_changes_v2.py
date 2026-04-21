"""Final rate_changes.xlsx rebuild with audit corrections + strict NAIC TOI codes.

Inputs:
    output/rate_changes.xlsx          — 23-row prior deliverable (column source)
    output/all_states_final.xlsx      — full filing metadata (TOI / sub-TOI codes)
    output/rate_change_audit.json     — PDF cross-check classifications
    output/strict_classification.json — strict NAIC TOI decisions
    output/effective_dates.json       — SERFF detail-page effective dates (overlay)

Outputs:
    output/rate_changes.xlsx (3 sheets: Rate Changes / Legend & Notes / Manual Review)
    output/rate_changes.csv

Corrections applied:
    1. 5 PDF-overall corrections (rate_effect_value updated, original_value preserved)
    2. 3 indicated-only rows (rate_effect_value nulled, original_value preserved)
    3. SFMA-134522369 split into 3 rows by Homeowners form
    4. New columns: toi_code, sub_toi_code, filing_component, rate_change_type,
       original_value, correction_note
    5. Legend references the NAIC Uniform P&C Product Coding Matrix
"""
from __future__ import annotations

import csv
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import OUTPUT_DIR

ALL_STATES = OUTPUT_DIR / "all_states_final.xlsx"
PRIOR = OUTPUT_DIR / "rate_changes.xlsx"
AUDIT_JSON = OUTPUT_DIR / "rate_change_audit.json"
EFFECTIVE_DATES = OUTPUT_DIR / "effective_dates.json"
PDF_ROOT = OUTPUT_DIR / "pdfs"

OUT_XLSX = OUTPUT_DIR / "rate_changes.xlsx"
OUT_CSV = OUTPUT_DIR / "rate_changes.csv"


# ---------- per-filing corrections (audit-driven, hand-verified snippets) ----------

# Each entry overrides rate_effect_value and sets rate_change_type.
PDF_CORRECTIONS = {
    "SFMA-134702926": {
        "new_value": -7.80,
        "rate_change_type": "overall_impact",
        "correction_note": (
            "SERFF value (+0.30%) matched the indicated rate level change for SFM. "
            "PDF (\"CO Filing Packet - SFM Updated.pdf\") states the proposed overall "
            "change is -7.80% for State Farm Mutual."
        ),
    },
    "SFMA-134872376": {
        "new_value": 9.90,
        "rate_change_type": "overall_impact",
        "correction_note": (
            "SERFF value (+0.00%) was placeholder. PDF (\"ID HO 2026 Filing.pdf\") "
            "states a statewide average 9.9% change for the Non-Tenant Homeowners form."
        ),
    },
    "ALSE-134538132": {
        "new_value": 8.30,
        "rate_change_type": "overall_impact",
        "correction_note": (
            "SERFF Requested Rate Effect (+26.60%) is the Allstate-bucket request. "
            "PDF (\"3. WA PPA AI R58865 Filing Memo.pdf\") states the proposed overall "
            "rate level change is 8.30% (after Rate Adjustment Factor offset)."
        ),
    },
    "ALSE-134416811": {
        "new_value": -13.70,
        "rate_change_type": "overall_impact",
        "correction_note": (
            "SERFF value (+0.00%) was placeholder. PDF "
            "(\"3. WA PPA ANAIC R58354 Filing Memo.pdf\") states a -13.7% overall rate "
            "level decrease referenced in the Rate and Rule Schedule."
        ),
    },
    # SFMA-134522369 is handled separately via the SPLIT_ROWS map below.
}

INDICATED_NULLS = {
    "LBPM-134551958": {
        "rate_change_type": "indicated",
        "correction_note": (
            "SERFF value (+7.20%) matched only a sub-coverage indicated rate change "
            "(Bodily Injury 7.2%). The PDF (\"CO_PL_AO_Exhibits_07-09-2025.pdf\") "
            "states the overall rate indication is -2.20%; the proposed overall is "
            "not extractable. Value nulled to avoid asserting an indicated number as "
            "a proposed rate change."
        ),
    },
    "SFMA-134532998": {
        "rate_change_type": "indicated",
        "correction_note": (
            "SERFF value (+0.00%) matched only the indicated rate level change. "
            "PDF (\"CO Filing Packet - SFM.pdf\") does not contain an extractable "
            "proposed overall change. Value nulled per the indicated-vs-proposed "
            "distinction."
        ),
    },
    "SFMA-134676753": {
        "rate_change_type": "indicated",
        "correction_note": (
            "SERFF value (-2.60%) matched only the indicated rate level change. "
            "PDF (\"ID Actuarial Memorandum - Amended.pdf\") does not contain an "
            "extractable proposed overall change. Value nulled per the indicated-vs-"
            "proposed distinction."
        ),
    },
}

# SFMA-134522369: replace single row with 3 rows, one per Homeowners form.
SPLIT_ROWS = {
    "SFMA-134522369": [
        {
            "filing_component": "Non-Tenant",
            "rate_value": 17.00,
            "sub_toi_code": "04.0003",
            "sub_toi_label": "04.0003 Owner Occupied Homeowners",
            "line_of_business": "04.0 Homeowners — 04.0003 Owner Occupied Homeowners",
            "correction_note": (
                "Split from SFMA-134522369 (which filed under combined sub-TOI "
                "04.0000). PDF (\"CO HO 2025 Filing.pdf\") states a statewide "
                "average 17.0% change for the Non-Tenant Homeowners policy form."
            ),
        },
        {
            "filing_component": "Renters",
            "rate_value": 8.00,
            "sub_toi_code": "04.0004",
            "sub_toi_label": "04.0004 Tenant Homeowners",
            "line_of_business": "04.0 Homeowners — 04.0004 Tenant Homeowners",
            "correction_note": (
                "Split from SFMA-134522369. PDF states a statewide average 8.0% "
                "change for the Renters policy form."
            ),
        },
        {
            "filing_component": "Condominium",
            "rate_value": 8.00,
            "sub_toi_code": "04.0001",
            "sub_toi_label": "04.0001 Condominium Homeowners",
            "line_of_business": "04.0 Homeowners — 04.0001 Condominium Homeowners",
            "correction_note": (
                "Split from SFMA-134522369. PDF states a statewide average 8.0% "
                "change for the Condominium Unitowners policy form."
            ),
        },
    ],
}


# ---------- column layout ----------

COLUMNS = [
    "state",
    "carrier",
    "company_name",
    "line_of_business",
    "toi_code",
    "sub_toi_code",
    "filing_component",
    "effective_date",
    "effective_date_source",
    "rate_effect_value",
    "rate_effect_source",
    "rate_change_type",
    "original_value",
    "correction_note",
    "current_avg_premium",
    "new_avg_premium",
    "serff_tracking_number",
    "filing_date",
    "disposition_status",
]

INACTIVE_TOKENS = ("WITHDRAWN", "DISAPPROVED")

# Ambiguous filings: the +0.00% rule/form filings with no PDF rate phrase to confirm.
# Default rate_change_type = "ambiguous" if filing isn't in the corrected/null/split sets
# AND the audit didn't classify it as overall_impact or indicated.


# ---------- helpers ----------

_TOI_RE = re.compile(r"^\s*(\d{2}\.\d)\b")
_SUB_RE = re.compile(r"^\s*(\d{2}\.\d{4})\b")


def extract_toi_code(s):
    if not s:
        return None
    m = _TOI_RE.match(str(s))
    return m.group(1) if m else None


def extract_sub_toi_code(s):
    if not s:
        return None
    m = _SUB_RE.match(str(s))
    return m.group(1) if m else None


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "1", "yes")


def _is_inactive(r: dict) -> bool:
    status = " ".join(
        str(r.get(k) or "").upper()
        for k in ("disposition_status", "state_status", "filing_status")
    )
    return any(tok in status for tok in INACTIVE_TOKENS)


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _fmt_pct(v) -> str:
    if v is None:
        return ""
    return f"{v:+.2f}%"


def _fmt_money(v) -> str:
    if v is None:
        return ""
    return f"${v:,.2f}"


def _fmt_date(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        return s


def _line_of_business(r: dict) -> str:
    toi = (r.get("type_of_insurance") or "").strip()
    sub = (r.get("sub_type_of_insurance") or "").strip()
    if toi and sub:
        return f"{toi} — {sub}"
    return sub or toi


def _date_key(iso: str) -> int:
    if not iso:
        return 0
    try:
        return int(datetime.fromisoformat(iso).strftime("%Y%m%d"))
    except ValueError:
        return 0


def _pick_rate_effect(r: dict) -> tuple[Optional[float], str]:
    overall = _to_float(r.get("overall_rate_effect"))
    requested = _to_float(r.get("requested_rate_effect"))
    approved = _to_float(r.get("approved_rate_effect"))
    populated = [(name, v) for name, v in
                 (("approved", approved), ("overall", overall), ("requested", requested))
                 if v is not None]
    if not populated:
        return None, ""
    return populated[0][1], populated[0][0]


def _load_all_states() -> dict[str, dict]:
    wb = openpyxl.load_workbook(ALL_STATES, read_only=True, data_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    by_serff = {}
    for r in rows[1:]:
        d = dict(zip(header, r))
        serff = d.get("serff_tracking_number")
        if serff:
            by_serff[serff] = d
    return by_serff


def _load_target_serffs() -> list[str]:
    """Re-derive the 23 target serff numbers from rate_changes.xlsx."""
    wb = openpyxl.load_workbook(PRIOR, read_only=True, data_only=True)
    ws = wb["Rate Changes"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out = []
    for r in rows[1:]:
        d = dict(zip(header, r))
        serff = d.get("serff_tracking_number")
        if serff:
            out.append(serff)
    return out


def _load_prior_overlays() -> dict[str, dict]:
    """Read existing rate_changes.xlsx to preserve effective_date + premium overlays
    that were written in the prior build (PDF-memo parsed)."""
    wb = openpyxl.load_workbook(PRIOR, read_only=True, data_only=True)
    ws = wb["Rate Changes"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out: dict[str, dict] = {}
    for r in rows[1:]:
        d = dict(zip(header, r))
        serff = d.get("serff_tracking_number")
        if not serff:
            continue
        out[serff] = {
            "effective_date": d.get("effective_date"),
            "effective_date_source": d.get("effective_date_source"),
            "current_avg_premium": d.get("current_avg_premium"),
            "new_avg_premium": d.get("new_avg_premium"),
        }
    return out


def _load_audit() -> dict[str, dict]:
    if not AUDIT_JSON.exists():
        return {}
    data = json.loads(AUDIT_JSON.read_text(encoding="utf-8"))
    return {a["serff_tracking_number"]: a for a in data}


def _load_effective_dates() -> dict[str, dict]:
    if not EFFECTIVE_DATES.exists():
        return {}
    return json.loads(EFFECTIVE_DATES.read_text(encoding="utf-8"))


# ---------- legend ----------

LEGEND_ROWS = [
    ("Column", "Meaning"),
    ("state", "US state of the filing (WA / ID / CO)."),
    ("carrier", "Target carrier searched (Allstate, Geico, Liberty, Progressive, State Farm, Travelers)."),
    ("company_name", "Specific underwriting company that filed (e.g. Allstate Fire & Casualty)."),
    ("line_of_business",
     "Type Of Insurance — Sub Type Of Insurance string from SERFF. For split rows, this "
     "reflects the per-form sub-TOI (not the original combined sub-TOI)."),
    ("toi_code",
     "NAIC Type Of Insurance code (e.g. 19.0 for Personal Auto, 04.0 for Homeowners). "
     "From the NAIC Uniform Property & Casualty Product Coding Matrix."),
    ("sub_toi_code",
     "NAIC Sub-TOI code (e.g. 19.0001 Private Passenger Auto, 04.0001 Condominium "
     "Homeowners). For split rows, the per-form sub-TOI is used (e.g. 04.0003, 04.0004, "
     "04.0001) rather than the carrier's filed combined code (04.0000)."),
    ("filing_component",
     "When a single SERFF filing covers multiple Homeowners forms with different rate "
     "impacts, this names the component (Non-Tenant / Renters / Condominium / etc.) "
     "that this row represents. Blank for single-form filings."),
    ("effective_date",
     "When the rate change takes / took effect. Blank where SERFF Filing Access does not "
     "expose this and the rate memo PDF doesn't include an extractable "
     "'Effective Date: MM/DD/YYYY' phrase."),
    ("effective_date_source",
     "How the effective_date was derived: 'serff:<bucket>' = SERFF detail-page label "
     "(approved/requested/proposed/plain). 'pdf_memo' = parsed from a downloaded rate "
     "memo PDF. Blank = no source available; needs manual review."),
    ("rate_effect_value",
     "Signed percentage change in rates (e.g. +11.90%). For 'overall_impact' rows this "
     "is the proposed weighted average overall rate level change. For 'indicated' rows "
     "this is BLANK (the SERFF value matched only the indicated/actuarial number, not "
     "the proposed change). For 'ambiguous' rows this is the carrier-filed SERFF value "
     "with no PDF cross-check."),
    ("rate_effect_source",
     "Which SERFF database field the value originated from. Priority used: "
     "approved > overall > requested. "
     "'approved'  = SERFF Approved Rate Impact field (regulator-approved per-company). "
     "'overall'   = SERFF Overall Rate Impact field (carrier-stated filing-level "
     "weighted average; for closed/approved filings = approved overall). "
     "'requested' = SERFF Requested Rate Impact field (carrier-filed, not yet approved). "
     "'pdf_memo'  = value sourced/corrected from the rate memo PDF, overriding SERFF. "
     "Blank = nulled out (indicated-only row)."),
    ("rate_change_type",
     "Classification of what the value represents, after PDF cross-check. "
     "'overall_impact' = proposed weighted average overall change (the headline number "
     "for the filing's effect — what the user wants). "
     "'indicated' = actuarially indicated rate change (NOT the proposed change). For "
     "rows of this type the rate_effect_value has been nulled to avoid asserting an "
     "indicated value as a rate change. "
     "'ambiguous' = no labeled rate phrase in the PDF; the SERFF value (typically 0.00%) "
     "is reported as-is. These are usually rule/form filings without a headline rate "
     "change number in the memo."),
    ("original_value",
     "The pre-correction SERFF value. Preserved when the value was changed by PDF cross-"
     "check (overall_impact corrections) or nulled (indicated rows), so nothing is lost."),
    ("correction_note",
     "Explains why the row was corrected, nulled, or split. References the specific PDF "
     "filename + key phrase that drove the change. Blank for unchanged rows."),
    ("current_avg_premium",
     "Average annual premium BEFORE the change. Blank where the PDF memo does not state "
     "a concrete dollar amount alongside the keyword 'average premium' (typical — most "
     "memos cite percent change only, not dollar baseline)."),
    ("new_avg_premium",
     "Average annual premium AFTER the change. Same caveat as current_avg_premium."),
    ("serff_tracking_number", "Unique SERFF identifier; can be used to look the filing up directly."),
    ("filing_date", "Submission date the carrier filed with the state regulator."),
    ("disposition_status", "Regulator decision (Approved / Pending / etc)."),
    ("", ""),
    ("METHODOLOGY", ""),
    ("Line-of-business filter",
     "Strict NAIC Uniform Property & Casualty Product Coding Matrix code matching. "
     "Included sub-TOIs: 19.0000 Personal Auto Combinations, 19.0001 Private Passenger "
     "Auto, 19.0002 Motorcycle, 19.0003 Recreational Vehicle, 19.0004 Other Personal "
     "Auto, 04.0000 Homeowners Sub-TOI Combinations, 04.0001 Condominium Homeowners, "
     "04.0002 Mobile Homeowners, 04.0003 Owner Occupied Homeowners, 04.0004 Tenant "
     "Homeowners, 04.0005 Other Homeowners. Explicitly excluded: 03.0000 Personal "
     "Farmowners, 17.0021/17.1021/17.2021 Personal Umbrella, 20.xxxx Commercial Auto, "
     "30.0000/30.1000 Homeowner-Auto Combinations / Dwelling Fire (none in dataset), "
     "33.0001 Other Personal Lines, all other TOI prefixes."),
    ("Rate-effect classification",
     "Each rate_effect_value was cross-checked against the downloaded rate memo PDF "
     "using labeled phrase patterns. Distinguishes proposed overall change ("
     "'overall rate change of X%' / 'statewide average X% change') from indicated rate "
     "change ('indicated rate level change ... is X%' / 'overall rate indication X%'). "
     "Where the SERFF value matched only an indicated phrase and no proposed-overall "
     "phrase, the value was nulled rather than reported as a rate change."),
    ("Form splits",
     "Where one SERFF filing covers multiple Homeowners forms (Owner Occupied / Tenant / "
     "Condominium) with different rate impacts, the row was split per form using the "
     "per-form sub-TOI code rather than the carrier's filed combined code (04.0000). "
     "filing_component identifies the form and correction_note documents the split."),
    ("", ""),
    ("LIMITATIONS", ""),
    ("Effective date coverage",
     "SERFF Filing Access (the public consumer portal at filingaccess.serff.com) does NOT "
     "expose Rate Data / per-company effective dates anywhere on the filing detail page — "
     "only Submission Date, Disposition Date, and State Status Last Changed are available. "
     "PDF rate memos sometimes contain effective dates but most use boilerplate phrasing "
     "without an extractable date. Blank effective_date is the honest state of the data. "
     "See the Manual Review sheet for the rows needing hand-fill, with PDF directory paths."),
    ("Premium dollar coverage",
     "Most filings disclose only percent change, not the dollar baseline. Conservative "
     "pattern matching (requires 'average premium' keyword within 40 characters of the "
     "dollar figure) produced 1/25 hits. Approving broader patterns risks false "
     "positives, so columns are left blank rather than guessed."),
    ("Ambiguous +0% rows",
     "11 rows are flagged rate_change_type=ambiguous because no labeled rate phrase was "
     "found in any PDF. All have rate_effect_value=+0.00% in SERFF, consistent with rule/"
     "form filings (clarifications, manual updates) that don't carry a headline rate "
     "change. They're retained because SERFF flagged them as Rate/Rule filings, but "
     "their values cannot be PDF-confirmed."),
    ("Authoritative reference",
     "https://content.naic.org/sites/default/files/inline-files/Property%20%26%20Casualty"
     "%20Product%20Coding%20Matrix.pdf "
     "— the NAIC Uniform Property & Casualty Product Coding Matrix is the source of "
     "truth for all TOI / sub-TOI codes used in classification."),
]


# ---------- workbook writers ----------

def _write_xlsx(rows: list[dict], path: Path, manual_review_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate Changes"
    ws.append(COLUMNS)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
    for idx, col in enumerate(COLUMNS, start=1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"

    ws_legend = wb.create_sheet("Legend & Notes")
    for i, (col, meaning) in enumerate(LEGEND_ROWS, start=1):
        ws_legend.cell(row=i, column=1, value=col)
        ws_legend.cell(row=i, column=2, value=meaning)
    for cell in ws_legend["A"]:
        cell.font = bold
    ws_legend.column_dimensions["A"].width = 28
    ws_legend.column_dimensions["B"].width = 120
    for row_cells in ws_legend.iter_rows(min_row=1, max_row=ws_legend.max_row,
                                         min_col=2, max_col=2):
        for c in row_cells:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws_mr = wb.create_sheet("Manual Review (effective_date)")
    mr_cols = [
        "state", "carrier", "company_name", "line_of_business",
        "sub_toi_code", "rate_effect_value", "rate_change_type",
        "serff_tracking_number", "filing_date",
        "disposition_status", "pdf_directory",
    ]
    ws_mr.append(mr_cols)
    for cell in ws_mr[1]:
        cell.font = bold
        cell.fill = fill
    for r in manual_review_rows:
        ws_mr.append([r.get(c, "") for c in mr_cols])
    for idx, col in enumerate(mr_cols, start=1):
        max_len = len(col)
        for row in ws_mr.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws_mr.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 90)
    ws_mr.freeze_panes = "A2"

    wb.save(path)


def _write_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLUMNS})


# ---------- main ----------

def main() -> int:
    by_serff = _load_all_states()
    target_serffs = _load_target_serffs()
    audit = _load_audit()
    serff_dates = _load_effective_dates()
    prior_overlays = _load_prior_overlays()
    print(f"[load] {len(target_serffs)} target serffs from prior rate_changes.xlsx")

    out_rows: list[dict] = []

    for serff in target_serffs:
        full = by_serff.get(serff)
        if not full:
            print(f"  ! {serff} not found in all_states_final — skipping")
            continue

        # Skip the row that's being split — it gets replaced by SPLIT_ROWS entries
        if serff in SPLIT_ROWS:
            for split in SPLIT_ROWS[serff]:
                out_rows.append(_build_row(full, audit, serff_dates,
                                           prior_overlays, split=split))
            continue

        out_rows.append(_build_row(full, audit, serff_dates, prior_overlays))

    # Sort: state asc, then by sub_toi_code for stable grouping, then effective_date desc.
    out_rows.sort(
        key=lambda x: (
            x["state"],
            x["sub_toi_code"] or "",
            -_date_key(x["effective_date"] or x["filing_date"]),
        )
    )

    # Manual review queue: rows with blank effective_date.
    manual_review_rows = []
    for r in out_rows:
        if r["effective_date"]:
            continue
        # Look up filing_id from full record for PDF dir
        full = by_serff.get(r["serff_tracking_number"])
        fid = str(full.get("filing_id") or "") if full else ""
        state = r["state"]
        pdf_dir = (PDF_ROOT / state / fid)
        manual_review_rows.append({
            **r,
            "pdf_directory": str(pdf_dir.resolve()) if pdf_dir.exists() else "(not downloaded)",
        })

    _write_xlsx(out_rows, OUT_XLSX, manual_review_rows)
    _write_csv(out_rows, OUT_CSV)

    # ---- Final report ----
    total = len(out_rows)
    print("\n" + "=" * 60)
    print("FINAL REPORT")
    print("=" * 60)
    print(f"Total rows: {total}")

    print("\nBreakdown by sub_toi_code:")
    sub_counts: dict[str, int] = {}
    for r in out_rows:
        k = r["sub_toi_code"] or "(missing)"
        sub_counts[k] = sub_counts.get(k, 0) + 1
    for k in sorted(sub_counts):
        print(f"  {k}  {sub_counts[k]}")

    print("\nBreakdown by rate_change_type:")
    type_counts: dict[str, int] = {}
    for r in out_rows:
        k = r["rate_change_type"] or "(blank)"
        type_counts[k] = type_counts.get(k, 0) + 1
    for k in sorted(type_counts):
        print(f"  {k:18s}  {type_counts[k]}")

    print("\nData gap report:")
    print(f"  rate_effect_value populated:    "
          f"{sum(1 for r in out_rows if r['rate_effect_value']):>2d}/{total}")
    print(f"  effective_date populated:       "
          f"{sum(1 for r in out_rows if r['effective_date']):>2d}/{total}")
    print(f"  current_avg_premium populated:  "
          f"{sum(1 for r in out_rows if r['current_avg_premium']):>2d}/{total}")
    print(f"  new_avg_premium populated:      "
          f"{sum(1 for r in out_rows if r['new_avg_premium']):>2d}/{total}")
    print(f"  correction_note populated:      "
          f"{sum(1 for r in out_rows if r['correction_note']):>2d}/{total}")

    print(f"\nFiles written:")
    print(f"  {OUT_XLSX}")
    print(f"  {OUT_CSV}")
    return 0


def _build_row(full: dict, audit: dict, serff_dates: dict,
               prior_overlays: dict, *, split: Optional[dict] = None) -> dict:
    """Build one output row from a source filing record.

    If `split` is provided, override sub_toi_code/line_of_business/rate_value/
    filing_component/correction_note from the split spec instead of using the
    source record's combined code.
    """
    serff = full.get("serff_tracking_number")
    state = full.get("state") or ""
    target_company = full.get("target_company") or ""
    company_name = full.get("company_name") or ""
    toi = full.get("type_of_insurance")
    sub = full.get("sub_type_of_insurance")
    toi_code = extract_toi_code(toi)
    sub_code = extract_sub_toi_code(sub)
    line = _line_of_business(full)

    # Default values (no correction)
    rate_value, rate_source = _pick_rate_effect(full)
    rate_change_type = "ambiguous"
    original_value = None
    correction_note = ""
    filing_component = ""

    # Apply audit-derived classification (from audit JSON) for the default state.
    audit_hit = audit.get(serff)
    if audit_hit:
        rct = audit_hit.get("rate_change_type")
        if rct == "overall_impact":
            rate_change_type = "overall_impact"
        elif rct == "indicated":
            rate_change_type = "indicated"
        else:
            rate_change_type = "ambiguous"

    # Apply hand-verified PDF corrections
    if serff in PDF_CORRECTIONS and split is None:
        c = PDF_CORRECTIONS[serff]
        original_value = rate_value
        rate_value = c["new_value"]
        rate_change_type = c["rate_change_type"]
        correction_note = c["correction_note"]
        rate_source = "pdf_memo"

    # Apply indicated nulls
    if serff in INDICATED_NULLS:
        c = INDICATED_NULLS[serff]
        original_value = rate_value
        rate_value = None
        rate_change_type = c["rate_change_type"]
        correction_note = c["correction_note"]
        rate_source = ""

    # Apply split-row overrides
    if split is not None:
        original_value, _ = _pick_rate_effect(full)
        rate_value = split["rate_value"]
        sub_code = split["sub_toi_code"]
        line = split["line_of_business"]
        filing_component = split["filing_component"]
        correction_note = split["correction_note"]
        rate_change_type = "overall_impact"
        rate_source = "pdf_memo"

    # Effective date overlay (SERFF detail-page first, then PDF — already overlayed
    # in prior build, but re-derive for safety here).
    serff_hit = serff_dates.get(serff or "", {})
    serff_eff = serff_hit.get("effective_date") if isinstance(serff_hit, dict) else None
    serff_src = serff_hit.get("source") if isinstance(serff_hit, dict) else None
    eff = None
    eff_source = ""
    if serff_eff:
        eff = serff_eff
        eff_source = f"serff:{serff_src}" if serff_src else "serff"

    # Overlay from prior rate_changes.xlsx (preserves PDF-extracted values)
    overlay = prior_overlays.get(serff or "", {})
    if not eff and overlay.get("effective_date"):
        eff = overlay["effective_date"]
        eff_source = overlay.get("effective_date_source") or ""
    current_premium = overlay.get("current_avg_premium") or ""
    new_premium = overlay.get("new_avg_premium") or ""

    return {
        "state": state,
        "carrier": target_company,
        "company_name": company_name,
        "line_of_business": line,
        "toi_code": toi_code or "",
        "sub_toi_code": sub_code or "",
        "filing_component": filing_component,
        "effective_date": _fmt_date(eff),
        "effective_date_source": eff_source,
        "rate_effect_value": _fmt_pct(rate_value),
        "rate_effect_source": rate_source,
        "rate_change_type": rate_change_type,
        "original_value": _fmt_pct(original_value) if original_value is not None else "",
        "correction_note": correction_note,
        "current_avg_premium": current_premium,
        "new_avg_premium": new_premium,
        "serff_tracking_number": serff or "",
        "filing_date": _fmt_date(full.get("submission_date")),
        "disposition_status": full.get("disposition_status") or full.get("state_status") or "",
    }


if __name__ == "__main__":
    raise SystemExit(main())
