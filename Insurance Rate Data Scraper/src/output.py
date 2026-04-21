"""Excel output for scraped filings.

Sheets written:
  - "Filings":                one row per filing, all scraped fields
  - "Unparseable PDFs":       manual review queue for filings where the
                              rate-effect parser returned nothing despite PDFs
                              being attached (excludes new-product launches,
                              which are a legitimate category)
  - "New Product Launches":   filings where the memo indicates a greenfield
                              product/company launch — correctly have no rate
                              change to extract
  - "Withdrawn/Disapproved":  filings with rate effects extracted whose
                              disposition is WITHDRAWN or DISAPPROVED — these
                              rate changes never took effect
  - "Target Lines Only":      filings whose TOI/Sub-TOI matches config
                              TARGET_LINES (personal auto + homeowners)
  - "Target Lines - Rate Effects": extracted rate effects within TARGET_LINES,
                              filtered to exclude WITHDRAWN/DISAPPROVED — this
                              is the core product output
  - "Out of Scope Lines":     everything else (commercial auto, umbrella, etc.)
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import TARGET_LINES
from .models import Filing

FILINGS_COLUMNS = [
    "state",
    "serff_tracking_number",
    "filing_id",
    "company_name",
    "target_company",
    "naic_codes",
    "product_name",
    "type_of_insurance",
    "sub_type_of_insurance",
    "filing_type",
    "filing_status",
    "submission_date",
    "disposition_date",
    "disposition_status",
    "state_status",
    "in_target_lines",
    "is_resubmission_of",
    "requested_rate_effect",
    "approved_rate_effect",
    "overall_rate_effect",
    "affected_policyholders",
    "written_premium_volume",
    "annual_premium_impact_dollars",
    "current_avg_premium",
    "proposed_avg_premium",
    "premium_change_dollars",
    "program_name",
    "filing_reason",
    "prior_approval",
    "pdfs",
    "pdf_parse_status",
    "pdf_parse_fields_found",
    "detail_url",
]

INACTIVE_DISPOSITIONS = {"WITHDRAWN", "DISAPPROVED"}

UNPARSEABLE_COLUMNS = [
    "state",
    "serff_tracking_number",
    "filing_id",
    "company_name",
    "filing_type",
    "submission_date",
    "pdf_parse_status",
    "pdf_count",
    "pdf_names",
    "detail_url",
]


def _is_unparseable(f: Filing) -> bool:
    """A filing needs manual review if PDFs were attached, no rate field
    extracted, and it isn't a known new-product launch.

    `overall_rate_effect == 0.0` (premium-neutral sentinel) counts as parsed.
    """
    has_pdfs = bool(f.pdfs)
    no_rate_extracted = (
        f.overall_rate_effect is None
        and f.requested_rate_effect is None
        and f.approved_rate_effect is None
    )
    if f.pdf_parse_status == "new_product_launch":
        return False
    return has_pdfs and no_rate_extracted


def _has_rate_effect(f: Filing) -> bool:
    return (
        f.overall_rate_effect is not None
        or f.requested_rate_effect is not None
        or f.approved_rate_effect is not None
    )


def _is_inactive_disposition(f: Filing) -> bool:
    status = (f.disposition_status or f.state_status or f.filing_status or "").upper()
    return any(tok in status for tok in INACTIVE_DISPOSITIONS)


def _matches_target_lines(f: Filing) -> bool:
    haystack = " ".join(
        s.lower() for s in (f.type_of_insurance or "", f.sub_type_of_insurance or "")
    )
    return any(line in haystack for line in TARGET_LINES)


def annotate_filings(filings: Iterable[Filing]) -> list[Filing]:
    """Populate derived flags (`in_target_lines`) on each Filing in-place."""
    out = list(filings)
    for f in out:
        f.in_target_lines = _matches_target_lines(f)
    return out


def write_excel(filings: Iterable[Filing], output_path: Path) -> Path:
    """Write filings to xlsx. Returns the output path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    filings_list = annotate_filings(filings)

    wb = Workbook()
    ws_filings = wb.active
    ws_filings.title = "Filings"
    _write_header(ws_filings, FILINGS_COLUMNS)

    ws_unparseable = wb.create_sheet("Unparseable PDFs")
    _write_header(ws_unparseable, UNPARSEABLE_COLUMNS)

    ws_launches = wb.create_sheet("New Product Launches")
    _write_header(ws_launches, UNPARSEABLE_COLUMNS)

    ws_withdrawn = wb.create_sheet("Withdrawn-Disapproved")
    _write_header(ws_withdrawn, FILINGS_COLUMNS)

    ws_target = wb.create_sheet("Target Lines Only")
    _write_header(ws_target, FILINGS_COLUMNS)

    ws_target_effects = wb.create_sheet("Target Lines - Rate Effects")
    _write_header(ws_target_effects, FILINGS_COLUMNS)

    ws_out_of_scope = wb.create_sheet("Out of Scope Lines")
    _write_header(ws_out_of_scope, FILINGS_COLUMNS)

    for f in filings_list:
        row = f.to_row()
        row_values = [row.get(c) for c in FILINGS_COLUMNS]
        ws_filings.append(row_values)

        if _is_unparseable(f):
            ws_unparseable.append(_review_row(f))

        if f.pdf_parse_status == "new_product_launch":
            ws_launches.append(_review_row(f))

        has_effect = _has_rate_effect(f)
        if has_effect and _is_inactive_disposition(f):
            ws_withdrawn.append(row_values)

        if f.in_target_lines:
            ws_target.append(row_values)
            if has_effect and not _is_inactive_disposition(f):
                ws_target_effects.append(row_values)
        else:
            ws_out_of_scope.append(row_values)

    for ws, cols in (
        (ws_filings, FILINGS_COLUMNS),
        (ws_unparseable, UNPARSEABLE_COLUMNS),
        (ws_launches, UNPARSEABLE_COLUMNS),
        (ws_withdrawn, FILINGS_COLUMNS),
        (ws_target, FILINGS_COLUMNS),
        (ws_target_effects, FILINGS_COLUMNS),
        (ws_out_of_scope, FILINGS_COLUMNS),
    ):
        _autosize(ws, cols)
        ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def _review_row(f: Filing) -> list:
    return [
        f.state,
        f.serff_tracking_number,
        f.filing_id,
        f.company_name,
        f.filing_type,
        f.submission_date.isoformat() if f.submission_date else None,
        f.pdf_parse_status,
        len(f.pdfs),
        ";".join(p.display_name for p in f.pdfs),
        f.detail_url,
    ]


def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill


def _autosize(ws, columns: list[str]) -> None:
    for idx, col in enumerate(columns, start=1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
